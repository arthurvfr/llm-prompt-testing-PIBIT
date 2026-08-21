"""
Experimento de ENGENHARIA DE PROMPT (ObjEsp2) por ablação.

Roda o MESMO modelo local sobre as MESMAS submissões variando apenas o prompt,
para medir quanto cada componente do prompt contribui para o desempenho.

    P0  Baseline ingênuo   -> só "seja um professor e dê feedback" + esquema JSON
    P2  + regras + testes  -> P0 + regras pedagógicas + casos de teste oficiais
    P3  + Logic-of-Thought -> P2 + as três fases de raciocínio (prompt atual)

A diferença P2-P0 isola o ganho das regras pedagógicas e da verificação por casos
de teste; a diferença P3-P2 isola o ganho do raciocínio estruturado (LoT).

IMPORTANTE: o gabarito (colunas K-M) NÃO muda entre as variantes — o erro real do
código independe do prompt. O script copia o gabarito da aba de referência, então
não é preciso reexecutar os códigos dos alunos.

A variante P3 já foi executada e vive na aba "Avaliacao_Local". Por isso o padrão
é rodar apenas P0 e P2 (100 chamadas). Use --variantes P3 só se quiser refazê-la.

Uso:
    python avaliar_prompts.py --testar                 # testa a conexão
    python avaliar_prompts.py --preparar               # só cria as abas
    python avaliar_prompts.py                          # roda P0 e P2 (padrão)
    python avaliar_prompts.py --variantes P0           # só uma variante
    python avaliar_prompts.py --variantes P0,P2,P3     # tudo, refazendo o P3
    python avaliar_prompts.py --variantes P0 --linhas 2,5,10-12
    python avaliar_prompts.py --variantes P0 --amostra 20   # amostra estratificada
    python avaliar_prompts.py --refazer --variantes P0      # sobrescreve o que já existe

O script salva a planilha a cada linha: se travar ou faltar energia, é só rodar de
novo que ele continua de onde parou (as linhas já preenchidas são puladas).

Requisitos: openpyxl (a chamada HTTP usa só a biblioteca padrão).
"""
import argparse
import json
import os
import re
import shutil
import urllib.error
import urllib.request
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLANILHA_FILE = os.path.join(RAIZ, "planilha_pesquisa.xlsx")
BACKUP_DIR = os.path.join(RAIZ, "backups")

LOCAL_URL = "http://localhost:1234/v1/chat/completions"
MODEL_LABEL = "Qwen3-Coder-30B-A3B-Instruct-UD-Q8_K_XL (local, llama.cpp)"
TIMEOUT_SEGUNDOS = 900
TEMPERATURA = 0.2
MAX_TOKENS = 2048

ABA_REFERENCIA = "Avaliacao"

ABA_POR_VARIANTE = {
    "P0": "Local_P0",
    "P2": "Local_P2",
    "P3": "Avaliacao_Local",
}
VARIANTES_PADRAO = ["P0", "P2"]

DESCRICAO_VARIANTE = {
    "P0": "Baseline ingênuo (sem regras, sem casos de teste, sem LoT)",
    "P2": "Regras pedagógicas + casos de teste (sem LoT)",
    "P3": "Completo: regras + casos de teste + Logic-of-Thought",
}

TIPO_SEM_ERRO = "Nenhum (código correto)"
LINHA_SEM_ERRO = "-"


banco_problemas = {
    "1132": "Escreva um programa que leia dois valores inteiros X e Y. Calcule e mostre a soma de todos os números não múltiplos de 13 entre X e Y, incluindo ambos.",
    "1153": "Ler um valor N (N < 13). Calcular e escrever seu respectivo fatorial (N!). Fatorial de N = N * (N-1) * (N-2) * (N-3) * ... * 1.",
    "1828": "Bazinga! Ler número de casos (N). Ler escolhas de Sheldon e Raj. Imprimir 'Caso #i: Bazinga!' se Sheldon vence, 'Raj trapaceou!' se Raj vence, 'De novo!' se empatar. Regras: tesoura corta papel, papel cobre pedra, pedra esmaga lagarto, lagarto envenena Spock, Spock esmaga tesoura, tesoura decapita lagarto, lagarto come papel, papel refuta Spock, Spock vaporiza pedra, pedra quebra tesoura.",
    "1873": "Ler número de casos (C). Ler escolhas de rajesh e sheldon. Imprimir o nome do vencedor ('rajesh' ou 'sheldon') ou 'empate'. Segue as mesmas regras de pedra-papel-tesoura-lagarto-Spock do problema 1828.",
}

casos_teste = {
    "1132": [
        ("100\n200", "13954", "caso padrão da amostra oficial"),
        ("200\n100", "13954", "X maior que Y — exige inverter os valores"),
        ("13\n13", "0", "intervalo unitário sendo múltiplo de 13"),
        ("1\n1", "1", "intervalo unitário não múltiplo"),
        ("0\n0", "0", "zero é múltiplo de 13"),
        ("1\n12", "78", "nenhum múltiplo dentro do intervalo"),
        ("12\n14", "26", "múltiplo exatamente no meio"),
        ("26\n26", "0", "múltiplo maior que 13"),
        ("-14\n-12", "-26", "números negativos"),
        ("1\n1000", "462462", "intervalo grande"),
    ],
    "1153": [
        ("4", "24", "amostra oficial"),
        ("1", "1", "menor valor comum"),
        ("2", "2", "valor pequeno"),
        ("7", "5040", "valor intermediário"),
        ("10", "3628800", "valor alto"),
        ("12", "479001600", "valor máximo permitido"),
    ],
    "1828": [
        (
            "3\ntesoura papel\npedra Spock\nlagarto lagarto",
            "Caso #1: Bazinga!\nCaso #2: Raj trapaceou!\nCaso #3: De novo!",
            "amostra oficial",
        ),
        (
            "1\nSpock pedra",
            "Caso #1: Bazinga!",
            "caso único — a numeração começa em 1",
        ),
        (
            "5\nSpock Spock\nSpock tesoura\ntesoura Spock\nSpock papel\npapel Spock",
            "Caso #1: De novo!\nCaso #2: Bazinga!\nCaso #3: Raj trapaceou!\n"
            "Caso #4: Raj trapaceou!\nCaso #5: Bazinga!",
            "Spock escrito com S maiúsculo, como vem na entrada deste problema",
        ),
        (
            "6\npedra tesoura\npedra lagarto\nlagarto papel\nlagarto Spock\n"
            "papel pedra\ntesoura lagarto",
            "Caso #1: Bazinga!\nCaso #2: Bazinga!\nCaso #3: Bazinga!\n"
            "Caso #4: Bazinga!\nCaso #5: Bazinga!\nCaso #6: Bazinga!",
            "Sheldon vence todas — detecta matriz de regras invertida",
        ),
    ],
    "1873": [
        (
            "6\npedra tesoura\nspock spock\npapel lagarto\nlagarto spock\n"
            "tesoura pedra\npapel spock",
            "rajesh\nempate\nsheldon\nrajesh\nsheldon\nrajesh",
            "casos variados",
        ),
        (
            "5\npedra pedra\npapel papel\ntesoura tesoura\nlagarto lagarto\nspock spock",
            "empate\nempate\nempate\nempate\nempate",
            "todos empates",
        ),
        (
            "4\nspock tesoura\nspock pedra\ntesoura papel\ntesoura lagarto",
            "rajesh\nrajesh\nrajesh\nrajesh",
            "rajesh vence todas — detecta matriz de regras invertida",
        ),
        (
            "4\ntesoura spock\npedra spock\npapel tesoura\nlagarto tesoura",
            "sheldon\nsheldon\nsheldon\nsheldon",
            "sheldon vence todas — detecta ordem de leitura invertida",
        ),
    ],
}


def _indentar(texto, espacos=4):
    prefixo = " " * espacos
    return "\n".join(prefixo + linha for linha in str(texto).split("\n"))


def formatar_casos_teste(problema_id):
    casos = casos_teste.get(str(problema_id))
    if not casos:
        return ""
    linhas = []
    for i, (entrada, saida, descricao) in enumerate(casos, 1):
        linhas.append(
            f"Caso {i} ({descricao})\n"
            f"  Entrada:\n{_indentar(entrada)}\n"
            f"  Saída esperada:\n{_indentar(saida)}"
        )
    return "\n\n".join(linhas)


# ---------------------------------------------------------------------------
# As três variantes de prompt
#
# Todas exigem JSON e os MESMOS cinco campos de saída (is_correct, linha_do_erro,
# tipo_de_erro, feedback_formativo, dica_acao). Manter a saída idêntica é o que
# permite pontuar as três com a mesma rubrica. A única diferença estrutural é o
# campo logic_of_thought, exclusivo do P3 — porque esse campo É o LoT.
# ---------------------------------------------------------------------------
def prompt_p0(problem_description, student_code, problema_id=None):
    """Baseline ingênuo: nenhuma instrução pedagógica, nenhum caso de teste."""
    return f"""
Você é um professor de programação. Analise o código do aluno abaixo e forneça um feedback.

[Descrição do Problema]
{problem_description}

[Código do Aluno]
{student_code}

Retorne ESTRITAMENTE no formato JSON abaixo:
{{
  "is_correct": <true ou false>,
  "linha_do_erro": <numero da linha do erro, ou null se não houver erro>,
  "tipo_de_erro": "<Sintaxe, Execução ou Lógica>",
  "feedback_formativo": "<Seu comentário sobre o código do aluno.>",
  "dica_acao": "<Sua orientação sobre o que o aluno deve fazer.>"
}}
"""


def prompt_p2(problem_description, student_code, problema_id=None):
    """P0 + regras pedagógicas + casos de teste. Sem as fases do LoT."""
    casos_formatados = formatar_casos_teste(problema_id) or (
        "(nenhum caso de teste cadastrado para este problema — avalie apenas pela descrição acima)"
    )
    return f"""
Você é um professor de ciência da computação experiente, especialista em pedagogia para iniciantes em programação.
Sua tarefa é analisar o código de um aluno para um problema específico e fornecer um feedback formativo.

Regras INEGOCIÁVEIS:
1. NUNCA forneça a solução correta. Isso significa: NÃO escreva o código corrigido e NÃO descreva o algoritmo em texto plano (Ex: É PROIBIDO dizer "adicione +1 no range" ou "inverta a ordem das variáveis" ou "coloque dois pontos no final do for").
2. Foco em UM ERRO POR VEZ (Dica Progressiva). Se houver erro de sintaxe, foque APENAS no erro de sintaxe. Não mencione erros de lógica até que o código do aluno consiga rodar.
3. Método Socrático Estrito: A sua dica de ação deve ser uma pergunta guiada ou um CASO DE TESTE (Ex: "Simule seu código no papel com X=20 e Y=10. Quantas vezes o laço executa?").
4. O código deve ser avaliado ESTRITAMENTE sob as regras de sintaxe da linguagem Python 3. Ignorar comentários iniciados com '#'.
5. Use os CASOS DE TESTE fornecidos como a fonte da verdade sobre o comportamento do código.
   Só marque is_correct como true se o código produzir a saída esperada em TODOS eles.
   Se algum caso falhar, o erro é real: não classifique o código como correto.
6. Ao citar um caso de teste na dica, você PODE mostrar a entrada e perguntar o que o
   código do aluno produz com ela. Você NÃO PODE revelar a saída esperada, nem dizer que
   a saída do aluno está diferente dela — isso entregaria a correção e viola a regra 1.

[Descrição do Problema]
{problem_description}

[Casos de Teste Oficiais]
Estes casos definem o comportamento correto. Simule o código do aluno mentalmente em cada
um deles antes de responder. Eles são para o SEU uso na verificação; as saídas esperadas
NUNCA devem aparecer no feedback_formativo nem na dica_acao.
{casos_formatados}

[Código do Aluno]
{student_code}

Retorne ESTRITAMENTE no formato JSON abaixo:
{{
  "is_correct": <true ou false>,
  "linha_do_erro": <numero_da_linha_do_erro>,
  "tipo_de_erro": "<Sintaxe, Execução ou Lógica>",
  "feedback_formativo": "<Um comentário curto e amigável dizendo vagamente onde a execução tropeça. Fale de APENAS UM erro.>",
  "dica_acao": "<Uma pergunta socrática ou um caso de teste para o aluno investigar. NUNCA diga o que está faltando.>",
  "casos_que_falham": [<lista com os números dos casos de teste que o código não passa; lista vazia se passar em todos>]
}}
"""


def prompt_p3(problem_description, student_code, problema_id=None):
    """Prompt completo, idêntico ao de avaliar_lote_local.py. NÃO alterar sem
    refazer a aba Avaliacao_Local — a comparação depende de serem o mesmo texto."""
    casos_formatados = formatar_casos_teste(problema_id) or (
        "(nenhum caso de teste cadastrado para este problema — avalie apenas pela descrição acima)"
    )
    return f"""
Você é um professor de ciência da computação experiente, especialista em pedagogia para iniciantes em programação.
Sua tarefa é analisar o código de um aluno para um problema específico e fornecer um feedback formativo, utilizando a metodologia "Logic-of-Thought" (LoT).

Regras INEGOCIÁVEIS:
1. NUNCA forneça a solução correta. Isso significa: NÃO escreva o código corrigido e NÃO descreva o algoritmo em texto plano (Ex: É PROIBIDO dizer "adicione +1 no range" ou "inverta a ordem das variáveis" ou "coloque dois pontos no final do for").
2. Foco em UM ERRO POR VEZ (Dica Progressiva). Se houver erro de sintaxe, foque APENAS no erro de sintaxe. Não mencione erros de lógica até que o código do aluno consiga rodar.
3. Método Socrático Estrito: A sua dica de ação deve ser uma pergunta guiada ou um CASO DE TESTE (Ex: "Simule seu código no papel com X=20 e Y=10. Quantas vezes o laço executa?").
4. O código deve ser avaliado ESTRITAMENTE sob as regras de sintaxe da linguagem Python 3. Ignorar comentários iniciados com '#'.
5. Use os CASOS DE TESTE fornecidos como a fonte da verdade sobre o comportamento do código.
   Só marque is_correct como true se o código produzir a saída esperada em TODOS eles.
   Se algum caso falhar, o erro é real: não classifique o código como correto.
6. Ao citar um caso de teste na dica, você PODE mostrar a entrada e perguntar o que o
   código do aluno produz com ela. Você NÃO PODE revelar a saída esperada, nem dizer que
   a saída do aluno está diferente dela — isso entregaria a correção e viola a regra 1.

[Descrição do Problema]
{problem_description}

[Casos de Teste Oficiais]
Estes casos definem o comportamento correto. Simule o código do aluno mentalmente em cada
um deles antes de responder. Eles são para o SEU uso na verificação; as saídas esperadas
NUNCA devem aparecer no feedback_formativo nem na dica_acao.
{casos_formatados}

[Código do Aluno]
{student_code}

Na fase_3_traducao, planeje a abordagem, mas não escreva a solução aqui também.
Retorne ESTRITAMENTE no formato JSON abaixo:
{{
  "logic_of_thought": {{
    "fase_1_extracao": "<Atue como um interpretador Python 3. Leia o código e identifique a árvore lógica que o aluno implementou.>",
    "fase_2_extensao": "<Percorra os casos de teste um a um dizendo, para cada um, qual saída o código do aluno produziria e se ela bate com a esperada. Liste então todos os erros encontrados e ELEJA APENAS O ERRO MAIS CRÍTICO para focar agora (Sintaxe > Erro de Execução > Erro de Lógica).>",
    "fase_3_traducao": "<Planeje a abordagem pedagógica para este ÚNICO erro. Prefira usar a entrada de um dos casos de teste que falharam: qual pergunta socrática fará o aluno perceber a falha sozinho?>"
  }},
  "is_correct": <true ou false>,
  "linha_do_erro": <numero_da_linha_do_erro_escolhido_na_fase_2>,
  "tipo_de_erro": "<Sintaxe, Execução ou Lógica>",
  "feedback_formativo": "<Um comentário curto e amigável dizendo vagamente onde a execução tropeça. Fale de APENAS UM erro.>",
  "dica_acao": "<Uma pergunta socrática ou um caso de teste para o aluno investigar. NUNCA diga o que está faltando.>",
  "casos_que_falham": [<lista com os números dos casos de teste que o código não passa; lista vazia se passar em todos>]
}}
"""


MONTADOR = {"P0": prompt_p0, "P2": prompt_p2, "P3": prompt_p3}



# Comunicação com o llama-server
def _http_post_json(url, payload, timeout):
    dados = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url, data=dados, headers={"Content-Type": "application/json"}, method="POST"
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def extrair_json(texto):
    """Isola o JSON da resposta, removendo <think>, cercas markdown e texto solto."""
    if texto is None:
        raise ValueError("Resposta vazia do servidor.")
    texto = re.sub(r"<think>.*?</think>", "", texto, flags=re.DOTALL | re.IGNORECASE)
    texto = re.sub(r"^.*?</think>", "", texto, flags=re.DOTALL | re.IGNORECASE)
    texto = re.sub(r"```(?:json)?", "", texto)
    texto = texto.strip()
    try:
        return json.loads(texto)
    except json.JSONDecodeError:
        inicio, fim = texto.find("{"), texto.rfind("}")
        if inicio == -1 or fim == -1 or fim <= inicio:
            raise ValueError(f"Não encontrei JSON na resposta: {texto[:300]}")
        return json.loads(texto[inicio : fim + 1])


def chamar_llm(variante, problem_description, student_code, problema_id=None):
    prompt = MONTADOR[variante](problem_description, student_code, problema_id)
    payload = {
        "messages": [{"role": "user", "content": prompt}],
        "temperature": TEMPERATURA,
        "max_tokens": MAX_TOKENS,
        "response_format": {"type": "json_object"},
        "chat_template_kwargs": {"enable_thinking": False},
    }
    try:
        resposta = _http_post_json(LOCAL_URL, payload, TIMEOUT_SEGUNDOS)
    except urllib.error.URLError as e:
        raise RuntimeError(
            f"Não consegui falar com o servidor em {LOCAL_URL}. "
            f"O llama-server está rodando? Detalhe: {e}"
        ) from e
    return extrair_json(resposta["choices"][0]["message"]["content"])


def testar_conexao():
    print(f"Testando {LOCAL_URL} ...")
    try:
        r = _http_post_json(
            LOCAL_URL,
            {
                "messages": [{"role": "user", "content": "Responda apenas: {\"ok\": true}"}],
                "temperature": 0,
                "max_tokens": 64,
                "response_format": {"type": "json_object"},
                "chat_template_kwargs": {"enable_thinking": False},
            },
            60,
        )
        print("Servidor respondeu.")
        print("  modelo:", r.get("model", "(não informado)"))
        return True
    except Exception as e:
        print(f"[ERRO] {e}")
        print("Confira se o iniciar_server.bat está rodando e se a porta é 1234.")
        return False


# Abas
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BASE_FONT = Font(name="Arial", size=10)
BOLD = Font(bold=True, name="Arial", size=10)
LLM_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
GOLD_FILL = PatternFill("solid", start_color="D9EAD3", end_color="D9EAD3")
SCORE_FILL = PatternFill("solid", start_color="D9D2E9", end_color="D9D2E9")
SEC_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def cabecalhos(variante):
    p = variante  # P0 / P2 / P3
    return [
        ("Problem", 10), ("Student", 8), ("Attempt", 8), ("Status_Original", 20),
        ("Codigo_Aluno", 55),
        (f"{p}_is_correct", 14), (f"{p}_linha_erro", 14), (f"{p}_tipo_erro", 14),
        (f"{p}_feedback_formativo", 40), (f"{p}_dica_acao", 40),
        ("Gold_tipo_erro", 14), ("Gold_linha_erro", 14), ("Gold_descricao_correcao", 35),
        ("Diagnostico_Correto", 16),
        ("Nao_Revela_Solucao_1a5", 16), ("Foco_Unico_Erro_1a5", 16),
        ("Qualidade_Socratica_1a5", 16), ("Clareza_1a5", 12), ("Tom_Motivacional_1a5", 16),
        ("Comentarios", 35),
    ]


def ultima_linha_de_dados(ws):
    ultima = 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value is not None and ws.cell(row, 5).value is not None:
            ultima = row
    return ultima


def preparar_aba(wb, variante, verbose=True):
    """Cria a aba da variante copiando A-E e o gabarito K-M da aba de referência.

    Nunca sobrescreve F-J nem O-S de uma aba existente: são os dados já coletados.
    """
    ws_ref = wb[ABA_REFERENCIA]
    ultima = ultima_linha_de_dados(ws_ref)
    nome = ABA_POR_VARIANTE[variante]

    nova = nome not in wb.sheetnames
    ws = wb.create_sheet(nome) if nova else wb[nome]

    if nova:
        for i, (titulo, largura) in enumerate(cabecalhos(variante), start=1):
            c = ws.cell(1, i, titulo)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = largura
        ws.freeze_panes = "F2"
        ws.row_dimensions[1].height = 30

    for row in range(2, ultima + 1):
        for col in range(1, 6):          # A-E: identificação e código
            ws.cell(row, col, ws_ref.cell(row, col).value)
        ws.cell(row, 5).alignment = WRAP
        for col in range(11, 14):        # K-M: gabarito (independe do prompt)
            ws.cell(row, col, ws_ref.cell(row, col).value)
            ws.cell(row, col).fill = GOLD_FILL
        ws.cell(row, 14, f'=IF(AND(H{row}<>"",K{row}<>""),IF(H{row}=K{row},1,0),"")')
        for col in range(6, 11):
            ws.cell(row, col).fill = LLM_FILL
        for col in range(15, 20):
            ws.cell(row, col).fill = SCORE_FILL
        for col in range(1, 21):
            ws.cell(row, col).font = BASE_FONT
            ws.cell(row, col).border = BORDER

    if verbose:
        print(f"Aba '{nome}' ({variante}) {'criada' if nova else 'já existia'}: "
              f"{ultima - 1} linhas, gabarito copiado de '{ABA_REFERENCIA}'.")
    return ultima


def preparar_comparativo_prompts(wb, ultima):
    nome = "Comparativo_Prompts"
    if nome in wb.sheetnames:
        del wb[nome]
    ws = wb.create_sheet(nome)

    ws.column_dimensions["A"].width = 46
    for col in "BCD":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    abas = {v: ABA_POR_VARIANTE[v] for v in ("P0", "P2", "P3")}
    presentes = [v for v in ("P0", "P2", "P3") if abas[v] in wb.sheetnames]

    titulos = ["Métrica"] + [f"{v} — {DESCRICAO_VARIANTE[v]}" for v in presentes]
    for i, t in enumerate(titulos, start=1):
        c = ws.cell(1, i, t)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 46

    metricas = [
        ("Casos avaliados", "COUNT", "N"),
        ("% Diagnóstico correto", "AVERAGE", "N"),
        ("Média - Não revela solução", "AVERAGE", "O"),
        ("Média - Foco em um único erro", "AVERAGE", "P"),
        ("Média - Qualidade socrática", "AVERAGE", "Q"),
        ("Média - Clareza", "AVERAGE", "R"),
        ("Média - Tom motivacional", "AVERAGE", "S"),
    ]
    r = 2
    for rotulo, func, col in metricas:
        ws.cell(r, 1, rotulo).font = BOLD
        for j, v in enumerate(presentes, start=2):
            f = f"{func}({abas[v]}!{col}2:{col}{ultima})"
            ws.cell(r, j, f"={f}" if func == "COUNT" else f'=IFERROR({f},"")').font = BASE_FONT
        r += 1

    r += 1
    ws.cell(r, 1, "Falsos positivos (viu erro em código correto)").font = BOLD
    ws.cell(r, 1).fill = SEC_FILL
    r += 1
    CORRETO = '"Nenhum (código correto)"'
    ws.cell(r, 1, "Submissões corretas no gabarito").font = BASE_FONT
    for j, v in enumerate(presentes, start=2):
        ws.cell(r, j, f'=COUNTIF({abas[v]}!K2:K{ultima},{CORRETO})').font = BASE_FONT
    base = r
    r += 1
    ws.cell(r, 1, "Falsos positivos").font = BASE_FONT
    for j, v in enumerate(presentes, start=2):
        ws.cell(r, j, f'=SUMPRODUCT(({abas[v]}!K2:K{ultima}={CORRETO})*'
                      f'({abas[v]}!H2:H{ultima}<>"")*({abas[v]}!H2:H{ultima}<>{CORRETO}))').font = BASE_FONT
    fp = r
    r += 1
    ws.cell(r, 1, "Taxa de falso positivo").font = BASE_FONT
    for j in range(2, len(presentes) + 2):
        L = get_column_letter(j)
        ws.cell(r, j, f'=IFERROR({L}{fp}/{L}{base},"")').font = BASE_FONT
    r += 1
    ws.cell(r, 1, "Falsos negativos (disse 'correto' em código com erro)").font = BASE_FONT
    for j, v in enumerate(presentes, start=2):
        ws.cell(r, j, f'=SUMPRODUCT(({abas[v]}!K2:K{ultima}<>{CORRETO})*'
                      f'({abas[v]}!K2:K{ultima}<>"")*({abas[v]}!H2:H{ultima}={CORRETO}))').font = BASE_FONT

    r += 2
    ws.cell(r, 1, "% Diagnóstico correto por problema").font = BOLD
    ws.cell(r, 1).fill = SEC_FILL
    r += 1
    ws.cell(r, 1, "Problema").font = BOLD
    for j, v in enumerate(presentes, start=2):
        ws.cell(r, j, v).font = BOLD
    for prob in (1132, 1153, 1828, 1873):
        r += 1
        ws.cell(r, 1, prob).font = BASE_FONT
        for j, v in enumerate(presentes, start=2):
            ws.cell(r, j, f'=IFERROR(AVERAGEIF({abas[v]}!A2:A{ultima},{prob},'
                          f'{abas[v]}!N2:N{ultima}),"")').font = BASE_FONT

    r += 2
    ws.cell(r, 1, "Ganho por componente do prompt").font = BOLD
    ws.cell(r, 1).fill = SEC_FILL
    if "P0" in presentes and "P2" in presentes:
        r += 1
        ws.cell(r, 1, "P2 - P0  (efeito das regras + casos de teste)").font = BASE_FONT
        ws.cell(r, 2, '=IFERROR(C3-B3,"")').font = BASE_FONT
    if "P2" in presentes and "P3" in presentes:
        r += 1
        col_p2 = get_column_letter(presentes.index("P2") + 2)
        col_p3 = get_column_letter(presentes.index("P3") + 2)
        ws.cell(r, 1, "P3 - P2  (efeito do Logic-of-Thought)").font = BASE_FONT
        ws.cell(r, 2, f'=IFERROR({col_p3}3-{col_p2}3,"")').font = BASE_FONT

    print(f"Aba '{nome}' atualizada com as variantes: {', '.join(presentes)}.")



# Seleção de linhas
def parse_linhas(texto):
    linhas = set()
    for parte in texto.replace(";", ",").split(","):
        parte = parte.strip()
        if not parte:
            continue
        if "-" in parte:
            inicio, _, fim = parte.partition("-")
            inicio, fim = inicio.strip(), fim.strip()
            if not (inicio.isdigit() and fim.isdigit()):
                raise ValueError(f"Intervalo inválido: '{parte}'")
            if int(inicio) > int(fim):
                raise ValueError(f"Intervalo invertido: '{parte}'")
            linhas.update(range(int(inicio), int(fim) + 1))
        else:
            if not parte.isdigit():
                raise ValueError(f"Linha inválida: '{parte}'")
            linhas.add(int(parte))
    if not linhas:
        raise ValueError("Nenhuma linha informada.")
    if min(linhas) < 2:
        raise ValueError("A linha 1 é o cabeçalho. Use linhas a partir da 2.")
    return sorted(linhas)


def amostra_estratificada(ws, ultima, n):
    """Escolhe n linhas mantendo a proporção de cada problema e priorizando
    diversidade de tipo de erro no gabarito. Determinística: mesma amostra
    em todas as variantes, o que é essencial para a comparação ser pareada."""
    por_problema = {}
    for row in range(2, ultima + 1):
        prob = ws.cell(row, 1).value
        if prob is None:
            continue
        por_problema.setdefault(prob, []).append(row)

    total = sum(len(v) for v in por_problema.values())
    escolhidas = []
    for prob in sorted(por_problema):
        linhas = por_problema[prob]
        cota = max(1, round(n * len(linhas) / total))

        # Agrupa por tipo de erro e vai pegando um de cada tipo em rodízio, para
        # que a cota não fique concentrada numa única categoria (o que aconteceria
        # ao simplesmente ordenar por tipo e fatiar).
        por_tipo = {}
        for r in linhas:
            por_tipo.setdefault(str(ws.cell(r, 11).value), []).append(r)
        for grupo in por_tipo.values():
            grupo.sort()

        selecionadas, tipos = [], sorted(por_tipo)
        i = 0
        while len(selecionadas) < cota and any(por_tipo[t] for t in tipos):
            t = tipos[i % len(tipos)]
            if por_tipo[t]:
                selecionadas.append(por_tipo[t].pop(0))
            i += 1
        escolhidas.extend(selecionadas)
    return sorted(set(escolhidas))[:n]


# ---------------------------------------------------------------------------
# Execução
# ---------------------------------------------------------------------------
def criar_backup():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    nome = os.path.splitext(os.path.basename(PLANILHA_FILE))[0]
    destino = os.path.join(BACKUP_DIR, f"{nome}_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(PLANILHA_FILE, destino)
    return destino


def normalizar_diagnostico(resultado):
    if resultado.get("is_correct") is True:
        return LINHA_SEM_ERRO, TIPO_SEM_ERRO
    linha = resultado.get("linha_do_erro")
    tipo = resultado.get("tipo_de_erro")
    if tipo is None:
        tipo = "Não identificado"
    if linha is None:
        linha = LINHA_SEM_ERRO
    return linha, tipo


def processar(ws, row, variante, sobrescrever, posicao=""):
    problem = ws.cell(row, 1).value
    student = ws.cell(row, 2).value
    attempt = ws.cell(row, 3).value
    code = ws.cell(row, 5).value

    if problem is None or code is None:
        return "pulado", 0.0
    descricao = banco_problemas.get(str(problem))
    if not descricao:
        print(f"[AVISO] Problema {problem} (linha {row}) não está no dicionário.")
        return "pulado", 0.0
    if ws.cell(row, 6).value is not None and not sobrescrever:
        return "pulado", 0.0

    print(f"{posicao}[{variante}] linha {row} (prob={problem} aluno={student} tent={attempt})...",
          end=" ", flush=True)
    inicio = datetime.now()
    try:
        resultado = chamar_llm(variante, descricao, code, str(problem))
        dur = (datetime.now() - inicio).total_seconds()
        linha_erro, tipo_erro = normalizar_diagnostico(resultado)
        ws.cell(row, 6, resultado.get("is_correct"))
        ws.cell(row, 7, linha_erro)
        ws.cell(row, 8, tipo_erro)
        ws.cell(row, 9, resultado.get("feedback_formativo"))
        ws.cell(row, 10, resultado.get("dica_acao"))
        print(f"OK ({dur:.0f}s)")
        return "ok", dur
    except Exception as e:
        dur = (datetime.now() - inicio).total_seconds()
        print(f"[ERRO] {e}")
        return "falha", dur


def main():
    ap = argparse.ArgumentParser(
        description="Experimento de ablação de prompts (P0/P2/P3) com LLM local.")
    ap.add_argument("--variantes", default=",".join(VARIANTES_PADRAO),
                    help="Variantes a rodar, separadas por vírgula. Padrão: P0,P2 "
                         "(P3 já está em Avaliacao_Local).")
    ap.add_argument("--linhas", help="Linhas específicas. Ex: 2,5,10-12")
    ap.add_argument("--amostra", type=int,
                    help="Roda só uma amostra estratificada de N linhas (mesma amostra "
                         "em todas as variantes).")
    ap.add_argument("--refazer", action="store_true",
                    help="Sobrescreve linhas já preenchidas.")
    ap.add_argument("--testar", action="store_true", help="Só testa a conexão.")
    ap.add_argument("--preparar", action="store_true", help="Só cria as abas, sem avaliar.")
    args = ap.parse_args()

    if args.testar:
        testar_conexao()
        return

    variantes = [v.strip().upper() for v in args.variantes.split(",") if v.strip()]
    invalidas = [v for v in variantes if v not in MONTADOR]
    if invalidas:
        print(f"[ERRO] Variante(s) desconhecida(s): {invalidas}. Use P0, P2 ou P3.")
        return

    if args.preparar:
        wb = openpyxl.load_workbook(PLANILHA_FILE)
        ultima = 0
        for v in variantes:
            ultima = preparar_aba(wb, v)
        preparar_comparativo_prompts(wb, ultima)
        wb.save(PLANILHA_FILE)
        print(f"Planilha salva: {PLANILHA_FILE}")
        return

    if "P3" in variantes and not args.refazer:
        print("[AVISO] P3 já foi executada e está na aba 'Avaliacao_Local'. "
              "As linhas preenchidas serão puladas (use --refazer para regerar).")

    if not testar_conexao():
        return
    print()

    if args.refazer:
        print(f"Backup salvo em {criar_backup()}")

    wb = openpyxl.load_workbook(PLANILHA_FILE)
    ws_ref = wb[ABA_REFERENCIA]
    ultima = ultima_linha_de_dados(ws_ref)

    if args.linhas:
        linhas = [r for r in parse_linhas(args.linhas) if r <= ultima]
    elif args.amostra:
        linhas = amostra_estratificada(ws_ref, ultima, args.amostra)
        print(f"Amostra estratificada de {len(linhas)} linhas: {linhas}\n")
    else:
        linhas = list(range(2, ultima + 1))

    for v in variantes:
        preparar_aba(wb, v)
    wb.save(PLANILHA_FILE)

    total_chamadas = len(linhas) * len(variantes)
    print(f"\n{len(variantes)} variante(s) x {len(linhas)} linha(s) = {total_chamadas} chamadas.\n")

    duracoes = []
    for v in variantes:
        ws = wb[ABA_POR_VARIANTE[v]]
        print(f"--- Variante {v}: {DESCRICAO_VARIANTE[v]} ---")
        ok = falhas = pulados = 0
        for i, row in enumerate(linhas):
            status, dur = processar(ws, row, v, args.refazer, f"[{i+1}/{len(linhas)}] ")
            if status == "pulado":
                pulados += 1
                continue
            duracoes.append(dur)
            ok += status == "ok"
            falhas += status == "falha"
            wb.save(PLANILHA_FILE)
            if len(duracoes) >= 3 and (i + 1) % 10 == 0:
                media = sum(duracoes) / len(duracoes)
                restantes = total_chamadas - len(duracoes) - pulados
                if restantes > 0:
                    fim = datetime.now() + timedelta(seconds=media * restantes)
                    print(f"    ... média {media:.0f}s/chamada, "
                          f"faltam ~{restantes} -> término por volta de {fim:%H:%M}")
        print(f"--- {v}: {ok} ok, {falhas} falha(s), {pulados} pulada(s) ---\n")

    preparar_comparativo_prompts(wb, ultima)
    wb.save(PLANILHA_FILE)
    print(f"Concluído. Resultados em {PLANILHA_FILE}.")
    print("Compare as variantes na aba 'Comparativo_Prompts'.")
    print("\nAs colunas O-S das abas novas ficam em branco para pontuação. "
          "Use o MESMO critério da rubrica nas três variantes.")


if __name__ == "__main__":
    main()
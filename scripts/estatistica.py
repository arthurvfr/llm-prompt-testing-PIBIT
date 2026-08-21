"""
Análise estatística pareada dos resultados do estudo.

Todas as condições experimentais (dois modelos, três formulações de prompt)
foram aplicadas às MESMAS submissões, de modo que os dados são pareados por
construção. São aplicados:

  - McNemar, em forma exata (binomial), às medidas binárias de acerto. O teste
    considera apenas os pares discordantes, isto é, as submissões em que uma
    condição acerta e a outra erra, e verifica se essas discordâncias se
    distribuem igualmente entre as duas condições.

  - Wilcoxon dos postos sinalizados às notas ordinais de qualidade pedagógica
    (escala de 1 a 5), por não pressupor normalidade.

Como cada família de comparações envolve múltiplos testes, aplica-se correção
de Holm-Bonferroni. Os valores de p corrigidos são os que devem ser reportados.

O acerto do tipo de erro é calculado de duas formas:
  - sobre todas as submissões, incluindo as corretas, caso em que acertar
    significa também reconhecer a ausência de erro;
  - restrito às submissões que contêm erro, isolando a capacidade de
    diagnóstico da capacidade de reconhecer código correto.
A comparação entre as duas revela se uma diferença observada decorre do
diagnóstico ou da especificidade.

Uso:
    python estatistica.py [caminho_da_planilha]

Requisitos: openpyxl, scipy
"""
import os
import sys

import openpyxl
from scipy.stats import binomtest, wilcoxon

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ARQUIVO_PADRAO = os.path.join(RAIZ, "planilha_pesquisa.xlsx")
PRIMEIRA_LINHA = 2
ULTIMA_LINHA = 51
CORRETO = "Nenhum (código correto)"

# rótulo curto -> nome da aba
CONDICOES = {
    "Gemini": "Gemini_P3_LoT",
    "P3": "Qwen_P3_LoT",
    "P0": "Qwen_P0_Baseline",
    "P2": "Qwen_P2_RegrasTestes",
}

# colunas da planilha (1-based)
COL_TIPO_MODELO = 8
COL_LINHA_MODELO = 7
COL_GOLD_TIPO = 11
COL_GOLD_LINHA = 12

DIMENSOES = [
    ("Não revela solução", 16),
    ("Foco em erro único", 17),
    ("Qualidade socrática", 18),
    ("Clareza", 19),
    ("Tom motivacional", 20),
]


def coletar(ws):
    """Extrai, por linha da planilha, os acertos e as notas de uma condição.

    Devolve quatro dicionários indexados pelo número da linha:
      tipo_todas  - 1/0 para acerto do tipo em todas as submissões
      tipo_erro   - 1/0 para acerto do tipo apenas nas submissões com erro
      linha       - 1/0 para acerto da linha apenas nas submissões com erro
      notas       - {dimensão: {linha: nota}}
    """
    tipo_todas, tipo_erro, linha = {}, {}, {}
    notas = {nome: {} for nome, _ in DIMENSOES}

    for r in range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1):
        gold_tipo = ws.cell(r, COL_GOLD_TIPO).value
        if gold_tipo in (None, ""):
            continue

        acertou_tipo = int(ws.cell(r, COL_TIPO_MODELO).value == gold_tipo)
        tipo_todas[r] = acertou_tipo

        gold_linha = ws.cell(r, COL_GOLD_LINHA).value
        if gold_tipo != CORRETO:
            tipo_erro[r] = acertou_tipo
            if gold_linha not in (None, ""):
                linha[r] = int(ws.cell(r, COL_LINHA_MODELO).value == gold_linha)

        for nome, col in DIMENSOES:
            valor = ws.cell(r, col).value
            if isinstance(valor, (int, float)):
                notas[nome][r] = valor

    return tipo_todas, tipo_erro, linha, notas


def mcnemar_exato(acertos_a, acertos_b):
    """McNemar exato sobre duas séries pareadas de acertos binários.

    Devolve (n_pares, discordâncias a favor de A, a favor de B, valor de p).
    """
    comuns = sorted(set(acertos_a) & set(acertos_b))
    so_a = sum(1 for r in comuns if acertos_a[r] == 1 and acertos_b[r] == 0)
    so_b = sum(1 for r in comuns if acertos_a[r] == 0 and acertos_b[r] == 1)
    discordantes = so_a + so_b
    p = binomtest(so_a, discordantes, 0.5).pvalue if discordantes else 1.0
    return len(comuns), so_a, so_b, p


def wilcoxon_pareado(notas_a, notas_b):
    """Wilcoxon sobre duas séries pareadas de notas.

    Devolve (n_pares, diferença das médias A-B, valor de p). Se as séries forem
    idênticas o teste não se aplica e p é fixado em 1.
    """
    comuns = sorted(set(notas_a) & set(notas_b))
    x = [notas_a[r] for r in comuns]
    y = [notas_b[r] for r in comuns]
    diferenca = sum(x) / len(x) - sum(y) / len(y)
    if all(u == v for u, v in zip(x, y)):
        return len(comuns), diferenca, 1.0
    p = wilcoxon(x, y, zero_method="wilcox").pvalue
    return len(comuns), diferenca, p


def holm_bonferroni(valores_p):
    """Correção de Holm-Bonferroni, preservando a ordem original de entrada."""
    ordem = sorted(range(len(valores_p)), key=lambda i: valores_p[i])
    m = len(valores_p)
    ajustados = [0.0] * m
    anterior = 0.0
    for posicao, i in enumerate(ordem):
        valor = min(1.0, max(anterior, (m - posicao) * valores_p[i]))
        ajustados[i] = valor
        anterior = valor
    return ajustados


def comparar(dados, pares, titulo):
    """Executa todos os testes de uma família de comparações e imprime a tabela."""
    print("=" * 92)
    print(titulo)
    print("=" * 92)

    resultados, valores_p = [], []
    for a, b in pares:
        tipo_a, erro_a, linha_a, notas_a = dados[a]
        tipo_b, erro_b, linha_b, notas_b = dados[b]

        for rotulo, serie_a, serie_b in [
            ("acerto do tipo (todas)", tipo_a, tipo_b),
            ("acerto do tipo (com erro)", erro_a, erro_b),
            ("acerto da linha", linha_a, linha_b),
        ]:
            n, so_a, so_b, p = mcnemar_exato(serie_a, serie_b)
            resultados.append((f"{a} x {b}", rotulo, "McNemar", n,
                               f"{so_a} x {so_b}", p))
            valores_p.append(p)

        for nome, _ in DIMENSOES:
            n, dif, p = wilcoxon_pareado(notas_a[nome], notas_b[nome])
            resultados.append((f"{a} x {b}", nome, "Wilcoxon", n,
                               f"{dif:+.3f}", p))
            valores_p.append(p)

    ajustados = holm_bonferroni(valores_p)

    print(f"{'comparação':14}{'medida':28}{'teste':10}{'n':>4}"
          f"{'discord./dif':>14}{'p':>10}{'p Holm':>10}")
    print("-" * 92)
    for (comp, medida, teste, n, detalhe, p), p_aj in zip(resultados, ajustados):
        marca = " *" if p_aj < 0.05 else ""
        print(f"{comp:14}{medida:28}{teste:10}{n:>4}{detalhe:>14}"
              f"{p:>10.4f}{p_aj:>10.4f}{marca}")
    print()


def main():
    caminho = sys.argv[1] if len(sys.argv) > 1 else ARQUIVO_PADRAO
    wb = openpyxl.load_workbook(caminho)

    faltando = [aba for aba in CONDICOES.values() if aba not in wb.sheetnames]
    if faltando:
        print(f"[ERRO] Abas ausentes na planilha: {faltando}")
        return

    dados = {rotulo: coletar(wb[aba]) for rotulo, aba in CONDICOES.items()}

    n_total = len(dados["Gemini"][0])
    n_erro = len(dados["Gemini"][1])
    print(f"Submissões avaliadas: {n_total} ({n_erro} com erro, "
          f"{n_total - n_erro} corretas)\n")

    comparar(dados, [("Gemini", "P3")],
             "EXPERIMENTO 1 — comparação entre modelos (formulação P3 fixa)")
    comparar(dados, [("P2", "P0"), ("P3", "P2"), ("P3", "P0")],
             "EXPERIMENTO 2 — ablação de prompts (modelo local fixo)")

    print("Leitura: em McNemar, 'discord./dif' traz as discordâncias a favor de "
          "cada condição,\nna ordem em que aparecem na comparação; em Wilcoxon, "
          "a diferença entre as médias.\n'*' indica p < 0,05 após correção de "
          "Holm-Bonferroni.")


if __name__ == "__main__":
    main()

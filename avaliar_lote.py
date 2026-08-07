"""
Processa as submissões da planilha planilha_pesquisa.xlsx chamando o Gemini
com o mesmo prompt (Logic-of-Thought) do script original, e grava as respostas
nas colunas Gemini_*.

Rode localmente (precisa de acesso à API do Gemini e da sua GEMINI_API_KEY em config.py):
    python avaliar_lote.py                  # menu interativo
    python avaliar_lote.py --refazer        # REFAZ todas as avaliações, sobrescrevendo
    python avaliar_lote.py --tudo           # só as linhas ainda não avaliadas
    python avaliar_lote.py --linhas 5,7,10-12     # apenas essas linhas (reavaliação)
    python avaliar_lote.py --linhas 5,7 --forcar  # idem, sobrescrevendo sem perguntar

Na reavaliação por linhas, toda linha que já tiver avaliação é exibida e o script
pergunta se deve sobrescrevê-la com o novo feedback. Antes de qualquer execução que
sobrescreva avaliações existentes, uma cópia da planilha é salva em backups/.

Requisitos: google-genai, openpyxl
"""
import argparse
import json
import os
import shutil
import time
from datetime import datetime

import openpyxl
from google import genai
from google.genai import types
from config import GEMINI_API_KEY

PLANILHA_FILE = "planilha_pesquisa.xlsx"
BACKUP_DIR = "backups"
MODEL = "gemini-flash-latest"
SLEEP_BETWEEN_CALLS = 15

# Rótulo usado na coluna Gold_tipo_erro quando não há erro. O Gemini precisa gravar
# exatamente o mesmo texto, senão a fórmula da coluna N (Diagnostico_Correto)
# fica em branco nas submissões corretas.
TIPO_SEM_ERRO = "Nenhum (código correto)"
LINHA_SEM_ERRO = "-"

client = genai.Client(api_key=GEMINI_API_KEY)

banco_problemas = {
    "1132": "Escreva um programa que leia dois valores inteiros X e Y. Calcule e mostre a soma de todos os números não múltiplos de 13 entre X e Y, incluindo ambos.",
    "1153": "Ler um valor N (N < 13). Calcular e escrever seu respectivo fatorial (N!). Fatorial de N = N * (N-1) * (N-2) * (N-3) * ... * 1.",
    "1828": "Bazinga! Ler número de casos (N). Ler escolhas de Sheldon e Raj. Imprimir 'Caso #i: Bazinga!' se Sheldon vence, 'Raj trapaceou!' se Raj vence, 'De novo!' se empatar. Regras: tesoura corta papel, papel cobre pedra, pedra esmaga lagarto, lagarto envenena Spock, Spock esmaga tesoura, tesoura decapita lagarto, lagarto come papel, papel refuta Spock, Spock vaporiza pedra, pedra quebra tesoura.",
    "1873": "Ler número de casos (C). Ler escolhas de rajesh e sheldon. Imprimir o nome do vencedor ('rajesh' ou 'sheldon') ou 'empate'. Segue as mesmas regras de pedra-papel-tesoura-lagarto-Spock do problema 1828.",
}


# ---------------------------------------------------------------------------
# Casos de teste usados para verificar o código antes de gerar o feedback.
# Cada item: (entrada, saida_esperada, o_que_testa)
# As saídas foram conferidas contra soluções de referência.
# ---------------------------------------------------------------------------
casos_teste = {
    "1132": [
        ("100\n200", "13954", "caso padrão da amostra oficial"),
        ("200\n100", "13954", "X maior que Y — exige inverter os valores"),
        ("13\n13", "0", "intervalo unitário sendo múltiplo de 13"),
        ("1\n1", "1", "intervalo unitário não múltiplo"),
        ("0\n0", "0", "zero é múltiplo de 13"),
        ("1\n12", "78", "nenhum múltiplo dentro do intervalo"),
        ("12\n14", "26", "múltiplo exatamente no meio"),
        ("26\n26", "0", "múltiplo maior que 13"),
        ("-14\n-12", "-26", "números negativos"),
        ("1\n1000", "462462", "intervalo grande"),
    ],
    "1153": [
        ("4", "24", "amostra oficial"),
        ("1", "1", "menor valor comum"),
        ("2", "2", "valor pequeno"),
        ("7", "5040", "valor intermediário"),
        ("10", "3628800", "valor alto"),
        ("12", "479001600", "valor máximo permitido"),
    ],
    "1828": [
        (
            "3\ntesoura papel\npedra Spock\nlagarto lagarto",
            "Caso #1: Bazinga!\nCaso #2: Raj trapaceou!\nCaso #3: De novo!",
            "amostra oficial",
        ),
        (
            "1\nSpock pedra",
            "Caso #1: Bazinga!",
            "caso único — a numeração começa em 1",
        ),
        (
            "5\nSpock Spock\nSpock tesoura\ntesoura Spock\nSpock papel\npapel Spock",
            "Caso #1: De novo!\nCaso #2: Bazinga!\nCaso #3: Raj trapaceou!\n"
            "Caso #4: Raj trapaceou!\nCaso #5: Bazinga!",
            "Spock escrito com S maiúsculo, como vem na entrada deste problema",
        ),
        (
            "6\npedra tesoura\npedra lagarto\nlagarto papel\nlagarto Spock\n"
            "papel pedra\ntesoura lagarto",
            "Caso #1: Bazinga!\nCaso #2: Bazinga!\nCaso #3: Bazinga!\n"
            "Caso #4: Bazinga!\nCaso #5: Bazinga!\nCaso #6: Bazinga!",
            "Sheldon vence todas — detecta matriz de regras invertida",
        ),
    ],
    "1873": [
        (
            "6\npedra tesoura\nspock spock\npapel lagarto\nlagarto spock\n"
            "tesoura pedra\npapel spock",
            "rajesh\nempate\nsheldon\nrajesh\nsheldon\nrajesh",
            "casos variados",
        ),
        (
            "5\npedra pedra\npapel papel\ntesoura tesoura\nlagarto lagarto\nspock spock",
            "empate\nempate\nempate\nempate\nempate",
            "todos empates",
        ),
        (
            "4\nspock tesoura\nspock pedra\ntesoura papel\ntesoura lagarto",
            "rajesh\nrajesh\nrajesh\nrajesh",
            "rajesh vence todas — detecta matriz de regras invertida",
        ),
        (
            "4\ntesoura spock\npedra spock\npapel tesoura\nlagarto tesoura",
            "sheldon\nsheldon\nsheldon\nsheldon",
            "sheldon vence todas — detecta ordem de leitura invertida",
        ),
    ],
}


def formatar_casos_teste(problema_id):
    """Monta o bloco de casos de teste que vai dentro do prompt.

    Devolve string vazia se o problema não tiver casos cadastrados, para que o
    prompt continue funcionando com problemas novos ainda sem bateria de testes.
    """
    casos = casos_teste.get(str(problema_id))
    if not casos:
        return ""

    linhas = []
    for i, (entrada, saida, descricao) in enumerate(casos, 1):
        linhas.append(
            f"Caso {i} ({descricao})\n"
            f"  Entrada:\n{_indentar(entrada)}\n"
            f"  Saída esperada:\n{_indentar(saida)}"
        )
    return "\n\n".join(linhas)


def _indentar(texto, espacos=4):
    prefixo = " " * espacos
    return "\n".join(prefixo + linha for linha in str(texto).split("\n"))


def montar_prompt(problem_description, student_code, problema_id=None):
    casos_teste_formatados = formatar_casos_teste(problema_id) or (
        "(nenhum caso de teste cadastrado para este problema — avalie apenas "
        "pela descrição acima)"
    )
    return f"""
Você é um professor de ciência da computação experiente, especialista em pedagogia para iniciantes em programação.
Sua tarefa é analisar o código de um aluno para um problema específico e fornecer um feedback formativo, utilizando a metodologia "Logic-of-Thought" (LoT).

Regras INEGOCIÁVEIS:
1. NUNCA forneça a solução correta. Isso significa: NÃO escreva o código corrigido e NÃO descreva o algoritmo em texto plano (Ex: É PROIBIDO dizer "adicione +1 no range" ou "inverta a ordem das variáveis" ou "coloque dois pontos no final do for").
2. Foco em UM ERRO POR VEZ (Dica Progressiva). Se houver erro de sintaxe, foque APENAS no erro de sintaxe. Não mencione erros de lógica até que o código do aluno consiga rodar.
3. Método Socrático Estrito: A sua dica de ação deve ser uma pergunta guiada ou um CASO DE TESTE (Ex: "Simule seu código no papel com X=20 e Y=10. Quantas vezes o laço executa?").
4. O código deve ser avaliado ESTRITAMENTE sob as regras de sintaxe da linguagem Python 3. Ignorar comentários iniciados com '#'.
5. Use os CASOS DE TESTE fornecidos como a fonte da verdade sobre o comportamento do código.
   Só marque is_correct como true se o código produzir a saída esperada em TODOS eles.
   Se algum caso falhar, o erro é real: não classifique o código como correto.
6. Ao citar um caso de teste na dica, você PODE mostrar a entrada e perguntar o que o
   código do aluno produz com ela. Você NÃO PODE revelar a saída esperada, nem dizer que
   a saída do aluno está diferente dela — isso entregaria a correção e viola a regra 1.

[Descrição do Problema]
{problem_description}

[Casos de Teste Oficiais]
Estes casos definem o comportamento correto. Simule o código do aluno mentalmente em cada
um deles antes de responder. Eles são para o SEU uso na verificação; as saídas esperadas
NUNCA devem aparecer no feedback_formativo nem na dica_acao.
{casos_teste_formatados}

[Código do Aluno]
{student_code}

Na fase_3_traducao, planeje a abordagem, mas não escreva a solução aqui também.
Retorne ESTRITAMENTE no formato JSON abaixo:
{{
  "logic_of_thought": {{
    "fase_1_extracao": "<Atue como um interpretador Python 3. Leia o código e identifique a árvore lógica que o aluno implementou.>",
    "fase_2_extensao": "<Percorra os casos de teste um a um dizendo, para cada um, qual saída o código do aluno produziria e se ela bate com a esperada. Liste então todos os erros encontrados e ELEJA APENAS O ERRO MAIS CRÍTICO para focar agora (Sintaxe > Erro de Execução > Erro de Lógica).>",
    "fase_3_traducao": "<Planeje a abordagem pedagógica para este ÚNICO erro. Prefira usar a entrada de um dos casos de teste que falharam: qual pergunta socrática fará o aluno perceber a falha sozinho?>"
  }},
  "is_correct": <true ou false>,
  "linha_do_erro": <numero_da_linha_do_erro_escolhido_na_fase_2>,
  "tipo_de_erro": "<Sintaxe, Execução ou Lógica>",
  "feedback_formativo": "<Um comentário curto e amigável dizendo vagamente onde a execução tropeça. Fale de APENAS UM erro.>",
  "dica_acao": "<Uma pergunta socrática ou um caso de teste para o aluno investigar. NUNCA diga o que está faltando.>",
  "casos_que_falham": [<lista com os números dos casos de teste que o código não passa; lista vazia se passar em todos>]
}}
"""


def chamar_gemini(problem_description, student_code, problema_id=None):
    prompt = montar_prompt(problem_description, student_code, problema_id)
    response = client.models.generate_content(
        model=MODEL,
        contents=prompt,
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            temperature=0.2,
        ),
    )
    return json.loads(response.text)


def parse_linhas(texto):
    """Converte "5,7,10-12" na lista ordenada de linhas [5, 7, 10, 11, 12].

    Levanta ValueError se o texto tiver algo que não seja número ou intervalo.
    """
    linhas = set()

    for parte in texto.replace(";", ",").split(","):
        parte = parte.strip()
        if not parte:
            continue

        if "-" in parte:
            inicio, _, fim = parte.partition("-")
            inicio, fim = inicio.strip(), fim.strip()
            if not (inicio.isdigit() and fim.isdigit()):
                raise ValueError(f"Intervalo inválido: '{parte}'")
            if int(inicio) > int(fim):
                raise ValueError(f"Intervalo invertido: '{parte}'")
            linhas.update(range(int(inicio), int(fim) + 1))
        else:
            if not parte.isdigit():
                raise ValueError(f"Linha inválida: '{parte}'")
            linhas.add(int(parte))

    if not linhas:
        raise ValueError("Nenhuma linha informada.")

    if min(linhas) < 2:
        raise ValueError("A linha 1 é o cabeçalho da planilha. Use linhas a partir da 2.")

    return sorted(linhas)


def escolher_modo_interativo():
    """Menu usado quando o script roda sem argumentos.

    Retorna (linhas_selecionadas, politica). linhas_selecionadas = None significa lote completo.
    """
    print("Como deseja avaliar?")
    print("  [1] Lote completo (apenas as linhas ainda não preenchidas)")
    print("  [2] Escolher uma ou mais linhas da planilha (reavaliação)")
    print("  [3] REFAZER todas as avaliações (sobrescreve tudo que já está preenchido)")

    while True:
        opcao = input("Opção [1/2/3]: ").strip()

        if opcao == "1":
            return None, {"modo": "nunca"}

        if opcao == "2":
            try:
                linhas = parse_linhas(input("Linhas (ex: 5,7,10-12): "))
            except ValueError as e:
                print(f"[ERRO] {e}")
                continue
            # Linhas já preenchidas serão confirmadas uma a uma na hora de gravar.
            return linhas, {"modo": "perguntar"}

        if opcao == "3":
            print("Todas as colunas Gemini_* serão regeradas. Um backup será salvo antes.")
            if input("Confirma? [s/N]: ").strip().lower() in ("s", "sim"):
                return None, {"modo": "sempre"}
            print("Cancelado.")
            continue

        print("[ERRO] Digite 1, 2 ou 3.")


def resumir(texto, limite=110):
    """Deixa o feedback atual em uma linha só, para caber na pergunta de sobrescrita."""
    texto = " ".join(str(texto).split())
    return texto if len(texto) <= limite else texto[: limite - 1] + "…"


def confirmar_sobrescrita(ws, row, politica):
    """Pergunta se a avaliação já existente na linha deve ser sobrescrita.

    A política é mutável: as respostas "todas"/"nenhuma" valem para as linhas seguintes.
    """
    if politica["modo"] == "sempre":
        return True
    if politica["modo"] == "nunca":
        return False

    print(f"\n  A linha {row} já tem avaliação:")
    print(f"    is_correct : {ws.cell(row, 6).value}")
    print(f"    tipo_erro  : {ws.cell(row, 8).value} (linha {ws.cell(row, 7).value})")
    print(f"    feedback   : {resumir(ws.cell(row, 9).value)}")

    while True:
        resposta = input("  Sobrescrever com o novo feedback? [s]im / [n]ão / [t]odas / [q] nenhuma: ").strip().lower()

        if resposta in ("s", "sim"):
            return True
        if resposta in ("n", "nao", "não"):
            return False
        if resposta in ("t", "todas"):
            politica["modo"] = "sempre"
            return True
        if resposta in ("q", "nenhuma"):
            politica["modo"] = "nunca"
            return False

        print("  [ERRO] Responda s, n, t ou q.")


def criar_backup():
    """Copia a planilha para backups/ antes de sobrescrever avaliações. Retorna o caminho."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    nome = os.path.splitext(os.path.basename(PLANILHA_FILE))[0]
    destino = os.path.join(BACKUP_DIR, f"{nome}_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(PLANILHA_FILE, destino)
    return destino


def normalizar_diagnostico(resultado):
    """Devolve (linha_do_erro, tipo_de_erro) prontos para a planilha.

    Quando o código está correto o modelo devolve null nesses dois campos, o que
    deixava G e H vazias e, por consequência, a fórmula da coluna N em branco.
    Aqui esses casos viram o mesmo rótulo usado no gabarito.
    """
    is_correct = resultado.get("is_correct") is True
    linha = resultado.get("linha_do_erro")
    tipo = resultado.get("tipo_de_erro")

    if is_correct:
        return LINHA_SEM_ERRO, TIPO_SEM_ERRO

    # Código incorreto, mas o modelo não apontou o tipo/linha: preenche o que der.
    if tipo is None:
        tipo = "Não identificado"
    if linha is None:
        linha = LINHA_SEM_ERRO

    return linha, tipo


def normalizar_planilha(ws):
    """Corrige linhas já avaliadas que ficaram com G/H vazias. Retorna quantas mudou."""
    corrigidas = 0

    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 6).value is None:  # linha ainda não avaliada
            continue
        if ws.cell(row, 8).value is not None:  # tipo já preenchido
            continue

        resultado = {
            "is_correct": ws.cell(row, 6).value,
            "linha_do_erro": ws.cell(row, 7).value,
            "tipo_de_erro": ws.cell(row, 8).value,
        }
        linha_erro, tipo_erro = normalizar_diagnostico(resultado)

        ws.cell(row, 7, linha_erro)
        ws.cell(row, 8, tipo_erro)
        corrigidas += 1
        print(f"Linha {row}: is_correct={resultado['is_correct']} -> tipo='{tipo_erro}', linha='{linha_erro}'")

    return corrigidas


def processar(ws, row, politica, posicao=""):
    """Avalia uma única linha da planilha. Retorna 'ok', 'pulado' ou 'falha'."""
    problem = ws.cell(row, 1).value
    student = ws.cell(row, 2).value
    attempt = ws.cell(row, 3).value
    code = ws.cell(row, 5).value  # Coluna E (Codigo_Aluno)

    # Ignorar linhas vazias no fim da planilha
    if problem is None or code is None:
        print(f"PULANDO linha {row}: sem problema ou sem código.")
        return "pulado"

    problem_description = banco_problemas.get(str(problem))
    if not problem_description:
        print(f"[AVISO] Problema {problem} (linha {row}) não encontrado no dicionário.")
        return "pulado"

    # Linha já preenchida: no lote completo é pulada; na reavaliação, o usuário decide.
    if ws.cell(row, 6).value is not None and not confirmar_sobrescrita(ws, row, politica):
        print(f"PULANDO linha {row}: problema={problem} aluno={student} tentativa={attempt} (Já preenchido)")
        return "pulado"

    print(f"{posicao}Processando linha {row}: problema={problem} aluno={student} tentativa={attempt}...", end=" ")

    try:
        resultado = chamar_gemini(problem_description, code, str(problem))
        linha_erro, tipo_erro = normalizar_diagnostico(resultado)

        # Gravando nas colunas Gemini (F=6, G=7, H=8, I=9, J=10)
        ws.cell(row, 6, resultado.get("is_correct"))
        ws.cell(row, 7, linha_erro)
        ws.cell(row, 8, tipo_erro)
        ws.cell(row, 9, resultado.get("feedback_formativo"))
        ws.cell(row, 10, resultado.get("dica_acao"))

        # Informativo apenas: os casos que falharam aparecem no console para
        # conferência, sem criar coluna nova na planilha.
        falhos = resultado.get("casos_que_falham") or []
        if falhos:
            print(f"OK! (falhou nos casos de teste {falhos})")
        else:
            print("OK!")
        return "ok"

    except Exception as e:
        print(f"[ERRO]: {e}")
        return "falha"


def main():
    parser = argparse.ArgumentParser(description="Avalia submissões da planilha com o Gemini.")
    parser.add_argument("--linhas", help="Linhas da planilha a avaliar. Ex: 5,7,10-12")
    parser.add_argument("--tudo", action="store_true",
                        help="Lote completo (só as não avaliadas), sem menu interativo.")
    parser.add_argument("--refazer", action="store_true",
                        help="Refaz TODAS as avaliações, sobrescrevendo as já preenchidas.")
    parser.add_argument("--forcar", action="store_true",
                        help="Sobrescreve linhas já preenchidas sem perguntar.")
    parser.add_argument("--normalizar", action="store_true",
                        help="Só preenche G/H das linhas já avaliadas que ficaram vazias (não chama a API).")
    args = parser.parse_args()

    if args.normalizar:
        wb_rubrica = openpyxl.load_workbook(PLANILHA_FILE)
        corrigidas = normalizar_planilha(wb_rubrica["Avaliacao"])
        if corrigidas:
            wb_rubrica.save(PLANILHA_FILE)
        print(f"\n{corrigidas} linha(s) normalizada(s) em {PLANILHA_FILE}.")
        return

    if args.refazer:
        linhas_escolhidas = None
        politica = {"modo": "sempre"}
    elif args.linhas:
        linhas_escolhidas = parse_linhas(args.linhas)
        politica = {"modo": "sempre" if args.forcar else "perguntar"}
    elif args.tudo:
        linhas_escolhidas = None
        politica = {"modo": "sempre" if args.forcar else "nunca"}
    else:
        linhas_escolhidas, politica = escolher_modo_interativo()

    # Quem pode sobrescrever avaliação existente leva backup antes de começar.
    if politica["modo"] != "nunca":
        print(f"Backup salvo em {criar_backup()}")

    wb_rubrica = openpyxl.load_workbook(PLANILHA_FILE)
    ws = wb_rubrica["Avaliacao"]

    if linhas_escolhidas is None:
        linhas = list(range(2, ws.max_row + 1))
        acao = "Refazendo TODAS as avaliações" if politica["modo"] == "sempre" else "Processamento em lote"
        print(f"\n{acao} (linhas 2 a {ws.max_row})...")
    else:
        fora = [r for r in linhas_escolhidas if r > ws.max_row]
        if fora:
            print(f"[AVISO] Ignorando linhas fora da planilha (máx. {ws.max_row}): {fora}")
        linhas = [r for r in linhas_escolhidas if r <= ws.max_row]
        if not linhas:
            print("Nenhuma linha válida para processar.")
            return
        print(f"\nIniciando processamento das linhas: {linhas}")

    processadas, falhas = 0, 0

    print(f"Tempo estimado: ~{len(linhas) * SLEEP_BETWEEN_CALLS // 60} min (pausa de "
          f"{SLEEP_BETWEEN_CALLS}s entre chamadas por causa do rate limit).\n")

    for i, row in enumerate(linhas):
        status = processar(ws, row, politica, f"[{i + 1}/{len(linhas)}] ")

        if status == "pulado":
            continue

        processadas += 1
        if status == "falha":
            falhas += 1

        # Salva o arquivo a cada iteração para garantir que não haja perda de dados
        wb_rubrica.save(PLANILHA_FILE)

        # Respeita o limite de taxa (Rate Limit) da API — desnecessário após a última linha
        if i < len(linhas) - 1:
            time.sleep(SLEEP_BETWEEN_CALLS)

    print(f"\nConcluído. {processadas} submissões processadas, {falhas} falhas.")
    print(f"Resultados finais gravados com segurança em {PLANILHA_FILE}.")

if __name__ == "__main__":
    main()
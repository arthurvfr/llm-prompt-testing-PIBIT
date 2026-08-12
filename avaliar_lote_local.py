"""
Versão do avaliar_lote.py que usa um LLM LOCAL servido pelo llama-server
(llama.cpp) em vez da API do Gemini.

O llama-server expõe uma API compatível com a da OpenAI, então basta um POST em
http://localhost:1234/v1/chat/completions. Não há rate limit: a pausa entre
chamadas é zero e o gargalo é só a velocidade da sua GPU.

Antes de rodar, suba o servidor com o seu iniciar_server.bat e confira que a
porta bate com LOCAL_URL abaixo.

    python avaliar_lote_local.py                  # menu interativo
    python avaliar_lote_local.py --refazer        # REFAZ todas, sobrescrevendo
    python avaliar_lote_local.py --tudo           # só as linhas ainda não avaliadas
    python avaliar_lote_local.py --linhas 5,7,10-12     # apenas essas linhas
    python avaliar_lote_local.py --linhas 5,7 --forcar  # idem, sem perguntar
    python avaliar_lote_local.py --testar         # só testa a conexão com o servidor
    python avaliar_lote_local.py --preparar       # só cria/atualiza as abas, sem avaliar

Os resultados vão para a aba "Avaliacao_Local" da MESMA planilha, preservando
as colunas de gabarito (K-M) copiadas da aba "Avaliacao". A aba "Comparativo"
calcula as métricas dos dois modelos lado a lado.

Requisitos: openpyxl (a chamada HTTP usa só a biblioteca padrão do Python).
"""
import argparse
import json
import os
import re
import shutil
import urllib.error
import urllib.request
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuração do servidor local
# ---------------------------------------------------------------------------
PLANILHA_FILE = "planilha_pesquisa.xlsx"
BACKUP_DIR = "backups"

LOCAL_URL = "http://localhost:1234/v1/chat/completions"
# Nome apenas informativo: o llama-server responde com o modelo que estiver carregado.
# Este rótulo é o que vai para a planilha e para o relatório.
MODEL_LABEL = "Qwen3-Coder-30B-A3B-Instruct-UD-Q8_K_XL (local, llama.cpp)"
TIMEOUT_SEGUNDOS = 900  # modelo local pode demorar; generoso de propósito
TEMPERATURA = 0.2
MAX_TOKENS = 2048

ABA_GEMINI = "Avaliacao"
ABA_LOCAL = "Avaliacao_Local"
ABA_COMPARATIVO = "Comparativo"

TIPO_SEM_ERRO = "Nenhum (código correto)"
LINHA_SEM_ERRO = "-"

# ---------------------------------------------------------------------------
# Banco de problemas e casos de teste (idênticos aos do avaliar_lote.py)
# ---------------------------------------------------------------------------
banco_problemas = {
    "1132": "Escreva um programa que leia dois valores inteiros X e Y. Calcule e mostre a soma de todos os números não múltiplos de 13 entre X e Y, incluindo ambos.",
    "1153": "Ler um valor N (N < 13). Calcular e escrever seu respectivo fatorial (N!). Fatorial de N = N * (N-1) * (N-2) * (N-3) * ... * 1.",
    "1828": "Bazinga! Ler número de casos (N). Ler escolhas de Sheldon e Raj. Imprimir 'Caso #i: Bazinga!' se Sheldon vence, 'Raj trapaceou!' se Raj vence, 'De novo!' se empatar. Regras: tesoura corta papel, papel cobre pedra, pedra esmaga lagarto, lagarto envenena Spock, Spock esmaga tesoura, tesoura decapita lagarto, lagarto come papel, papel refuta Spock, Spock vaporiza pedra, pedra quebra tesoura.",
    "1873": "Ler número de casos (C). Ler escolhas de rajesh e sheldon. Imprimir o nome do vencedor ('rajesh' ou 'sheldon') ou 'empate'. Segue as mesmas regras de pedra-papel-tesoura-lagarto-Spock do problema 1828.",
}

casos_teste = {
    "1132": [
        ("100\n200", "13954", "caso padrão da amostra oficial"),
        ("200\n100", "13954", "X maior que Y — exige inverter os valores"),
        ("13\n13", "0", "intervalo unitário sendo múltiplo de 13"),
        ("1\n1", "1", "intervalo unitário não múltiplo"),
        ("0\n0", "0", "zero é múltiplo de 13"),
        ("1\n12", "78", "nenhum múltiplo dentro do intervalo"),
        ("12\n14", "26", "múltiplo exatamente no meio"),
        ("26\n26", "0", "múltiplo maior que 13"),
        ("-14\n-12", "-26", "números negativos"),
        ("1\n1000", "462462", "intervalo grande"),
    ],
    "1153": [
        ("4", "24", "amostra oficial"),
        ("1", "1", "menor valor comum"),
        ("2", "2", "valor pequeno"),
        ("7", "5040", "valor intermediário"),
        ("10", "3628800", "valor alto"),
        ("12", "479001600", "valor máximo permitido"),
    ],
    "1828": [
        (
            "3\ntesoura papel\npedra Spock\nlagarto lagarto",
            "Caso #1: Bazinga!\nCaso #2: Raj trapaceou!\nCaso #3: De novo!",
            "amostra oficial",
        ),
        (
            "1\nSpock pedra",
            "Caso #1: Bazinga!",
            "caso único — a numeração começa em 1",
        ),
        (
            "5\nSpock Spock\nSpock tesoura\ntesoura Spock\nSpock papel\npapel Spock",
            "Caso #1: De novo!\nCaso #2: Bazinga!\nCaso #3: Raj trapaceou!\n"
            "Caso #4: Raj trapaceou!\nCaso #5: Bazinga!",
            "Spock escrito com S maiúsculo, como vem na entrada deste problema",
        ),
        (
            "6\npedra tesoura\npedra lagarto\nlagarto papel\nlagarto Spock\n"
            "papel pedra\ntesoura lagarto",
            "Caso #1: Bazinga!\nCaso #2: Bazinga!\nCaso #3: Bazinga!\n"
            "Caso #4: Bazinga!\nCaso #5: Bazinga!\nCaso #6: Bazinga!",
            "Sheldon vence todas — detecta matriz de regras invertida",
        ),
    ],
    "1873": [
        (
            "6\npedra tesoura\nspock spock\npapel lagarto\nlagarto spock\n"
            "tesoura pedra\npapel spock",
            "rajesh\nempate\nsheldon\nrajesh\nsheldon\nrajesh",
            "casos variados",
        ),
        (
            "5\npedra pedra\npapel papel\ntesoura tesoura\nlagarto lagarto\nspock spock",
            "empate\nempate\nempate\nempate\nempate",
            "todos empates",
        ),
        (
            "4\nspock tesoura\nspock pedra\ntesoura papel\ntesoura lagarto",
            "rajesh\nrajesh\nrajesh\nrajesh",
            "rajesh vence todas — detecta matriz de regras invertida",
        ),
        (
            "4\ntesoura spock\npedra spock\npapel tesoura\nlagarto tesoura",
            "sheldon\nsheldon\nsheldon\nsheldon",
            "sheldon vence todas — detecta ordem de leitura invertida",
        ),
    ],
}


def _indentar(texto, espacos=4):
    prefixo = " " * espacos
    return "\n".join(prefixo + linha for linha in str(texto).split("\n"))


def formatar_casos_teste(problema_id):
    casos = casos_teste.get(str(problema_id))
    if not casos:
        return ""
    linhas = []
    for i, (entrada, saida, descricao) in enumerate(casos, 1):
        linhas.append(
            f"Caso {i} ({descricao})\n"
            f"  Entrada:\n{_indentar(entrada)}\n"
            f"  Saída esperada:\n{_indentar(saida)}"
        )
    return "\n\n".join(linhas)


def montar_prompt(problem_description, student_code, problema_id=None):
    """Mesmo prompt do avaliar_lote.py — manter idêntico é o que torna a
    comparação entre os modelos válida. Não altere aqui sem alterar lá."""
    casos_teste_formatados = formatar_casos_teste(problema_id) or (
        "(nenhum caso de teste cadastrado para este problema — avalie apenas "
        "pela descrição acima)"
    )
    return f"""
Você é um professor de ciência da computação experiente, especialista em pedagogia para iniciantes em programação.
Sua tarefa é analisar o código de um aluno para um problema específico e fornecer um feedback formativo, utilizando a metodologia "Logic-of-Thought" (LoT).

Regras INEGOCIÁVEIS:
1. NUNCA forneça a solução correta. Isso significa: NÃO escreva o código corrigido e NÃO descreva o algoritmo em texto plano (Ex: É PROIBIDO dizer "adicione +1 no range" ou "inverta a ordem das variáveis" ou "coloque dois pontos no final do for").
2. Foco em UM ERRO POR VEZ (Dica Progressiva). Se houver erro de sintaxe, foque APENAS no erro de sintaxe. Não mencione erros de lógica até que o código do aluno consiga rodar.
3. Método Socrático Estrito: A sua dica de ação deve ser uma pergunta guiada ou um CASO DE TESTE (Ex: "Simule seu código no papel com X=20 e Y=10. Quantas vezes o laço executa?").
4. O código deve ser avaliado ESTRITAMENTE sob as regras de sintaxe da linguagem Python 3. Ignorar comentários iniciados com '#'.
5. Use os CASOS DE TESTE fornecidos como a fonte da verdade sobre o comportamento do código.
   Só marque is_correct como true se o código produzir a saída esperada em TODOS eles.
   Se algum caso falhar, o erro é real: não classifique o código como correto.
6. Ao citar um caso de teste na dica, você PODE mostrar a entrada e perguntar o que o
   código do aluno produz com ela. Você NÃO PODE revelar a saída esperada, nem dizer que
   a saída do aluno está diferente dela — isso entregaria a correção e viola a regra 1.

[Descrição do Problema]
{problem_description}

[Casos de Teste Oficiais]
Estes casos definem o comportamento correto. Simule o código do aluno mentalmente em cada
um deles antes de responder. Eles são para o SEU uso na verificação; as saídas esperadas
NUNCA devem aparecer no feedback_formativo nem na dica_acao.
{casos_teste_formatados}

[Código do Aluno]
{student_code}

Na fase_3_traducao, planeje a abordagem, mas não escreva a solução aqui também.
Retorne ESTRITAMENTE no formato JSON abaixo:
{{
  "logic_of_thought": {{
    "fase_1_extracao": "<Atue como um interpretador Python 3. Leia o código e identifique a árvore lógica que o aluno implementou.>",
    "fase_2_extensao": "<Percorra os casos de teste um a um dizendo, para cada um, qual saída o código do aluno produziria e se ela bate com a esperada. Liste então todos os erros encontrados e ELEJA APENAS O ERRO MAIS CRÍTICO para focar agora (Sintaxe > Erro de Execução > Erro de Lógica).>",
    "fase_3_traducao": "<Planeje a abordagem pedagógica para este ÚNICO erro. Prefira usar a entrada de um dos casos de teste que falharam: qual pergunta socrática fará o aluno perceber a falha sozinho?>"
  }},
  "is_correct": <true ou false>,
  "linha_do_erro": <numero_da_linha_do_erro_escolhido_na_fase_2>,
  "tipo_de_erro": "<Sintaxe, Execução ou Lógica>",
  "feedback_formativo": "<Um comentário curto e amigável dizendo vagamente onde a execução tropeça. Fale de APENAS UM erro.>",
  "dica_acao": "<Uma pergunta socrática ou um caso de teste para o aluno investigar. NUNCA diga o que está faltando.>",
  "casos_que_falham": [<lista com os números dos casos de teste que o código não passa; lista vazia se passar em todos>]
}}
"""


# ---------------------------------------------------------------------------
# Comunicação com o llama-server
# ---------------------------------------------------------------------------
def _http_post_json(url, payload, timeout):
    dados = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url, data=dados, headers={"Content-Type": "application/json"}, method="POST"
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def extrair_json(texto):
    """Isola o objeto JSON da resposta do modelo.

    Modelos locais costumam sujar a saída de três formas:
      - bloco de raciocínio <think>...</think> (caso do Qwen3);
      - cercas de markdown ```json ... ```;
      - texto solto antes/depois do objeto.
    Aqui todas são removidas antes do json.loads.
    """
    if texto is None:
        raise ValueError("Resposta vazia do servidor.")

    # 1) remove blocos de raciocínio
    texto = re.sub(r"<think>.*?</think>", "", texto, flags=re.DOTALL | re.IGNORECASE)
    # <think> sem fechamento (resposta truncada)
    texto = re.sub(r"^.*?</think>", "", texto, flags=re.DOTALL | re.IGNORECASE)

    # 2) remove cercas de markdown
    texto = re.sub(r"```(?:json)?", "", texto)

    texto = texto.strip()

    # 3) tenta direto; se falhar, recorta do primeiro '{' ao último '}'
    try:
        return json.loads(texto)
    except json.JSONDecodeError:
        inicio, fim = texto.find("{"), texto.rfind("}")
        if inicio == -1 or fim == -1 or fim <= inicio:
            raise ValueError(f"Não encontrei JSON na resposta: {texto[:300]}")
        return json.loads(texto[inicio : fim + 1])


def chamar_llm_local(problem_description, student_code, problema_id=None):
    prompt = montar_prompt(problem_description, student_code, problema_id)
    payload = {
        "messages": [{"role": "user", "content": prompt}],
        "temperature": TEMPERATURA,
        "max_tokens": MAX_TOKENS,
        # llama.cpp aceita o response_format da OpenAI e força saída JSON válida
        "response_format": {"type": "json_object"},
        # desliga o modo "thinking" do Qwen3; servidores que não conhecem a
        # chave simplesmente a ignoram
        "chat_template_kwargs": {"enable_thinking": False},
    }
    try:
        resposta = _http_post_json(LOCAL_URL, payload, TIMEOUT_SEGUNDOS)
    except urllib.error.URLError as e:
        raise RuntimeError(
            f"Não consegui falar com o servidor em {LOCAL_URL}. "
            f"O llama-server está rodando? Detalhe: {e}"
        ) from e

    conteudo = resposta["choices"][0]["message"]["content"]
    return extrair_json(conteudo)


def testar_conexao():
    print(f"Testando {LOCAL_URL} ...")
    try:
        r = _http_post_json(
            LOCAL_URL,
            {
                "messages": [{"role": "user", "content": "Responda apenas: {\"ok\": true}"}],
                "temperature": 0,
                "max_tokens": 64,
                "response_format": {"type": "json_object"},
                "chat_template_kwargs": {"enable_thinking": False},
            },
            60,
        )
        print("Servidor respondeu.")
        print("  modelo:", r.get("model", "(não informado)"))
        print("  conteúdo:", r["choices"][0]["message"]["content"][:200])
        return True
    except Exception as e:
        print(f"[ERRO] {e}")
        print("Confira se o iniciar_server.bat está rodando e se a porta é 1234.")
        return False


# ---------------------------------------------------------------------------
# Preparação das abas
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BASE_FONT = Font(name="Arial", size=10)
LLM_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
GOLD_FILL = PatternFill("solid", start_color="D9EAD3", end_color="D9EAD3")
SCORE_FILL = PatternFill("solid", start_color="D9D2E9", end_color="D9D2E9")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CABECALHOS_LOCAL = [
    ("Problem", 10), ("Student", 8), ("Attempt", 8), ("Status_Original", 20),
    ("Codigo_Aluno", 55),
    ("Local_is_correct", 14), ("Local_linha_erro", 14), ("Local_tipo_erro", 14),
    ("Local_feedback_formativo", 40), ("Local_dica_acao", 40),
    ("Gold_tipo_erro", 14), ("Gold_linha_erro", 14), ("Gold_descricao_correcao", 35),
    ("Diagnostico_Correto", 16),
    ("Nao_Revela_Solucao_1a5", 16), ("Foco_Unico_Erro_1a5", 16),
    ("Qualidade_Socratica_1a5", 16), ("Clareza_1a5", 12), ("Tom_Motivacional_1a5", 16),
    ("Comentarios", 35),
]


def ultima_linha_de_dados(ws):
    ultima = 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value is not None and ws.cell(row, 5).value is not None:
            ultima = row
    return ultima


def preparar_abas(wb, verbose=True):
    """Cria (ou atualiza) a aba local copiando A-E e o gabarito K-M da aba Gemini.

    Nunca sobrescreve as colunas F-J e O-S de uma aba local já existente: elas
    guardam as avaliações e as notas que você já preencheu.
    """
    ws_gem = wb[ABA_GEMINI]
    ultima = ultima_linha_de_dados(ws_gem)

    nova = ABA_LOCAL not in wb.sheetnames
    ws = wb.create_sheet(ABA_LOCAL) if nova else wb[ABA_LOCAL]

    if nova:
        for i, (nome, largura) in enumerate(CABECALHOS_LOCAL, start=1):
            c = ws.cell(1, i, nome)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = largura
        ws.freeze_panes = "F2"
        ws.row_dimensions[1].height = 30

    for row in range(2, ultima + 1):
        # A-E: identificação e código (sempre sincronizados com a aba Gemini)
        for col in range(1, 6):
            ws.cell(row, col, ws_gem.cell(row, col).value)
        ws.cell(row, 5).alignment = WRAP

        # K-M: gabarito, copiado — é independente do modelo
        for col in range(11, 14):
            ws.cell(row, col, ws_gem.cell(row, col).value)
            ws.cell(row, col).fill = GOLD_FILL

        # N: fórmula de diagnóstico apontando para esta própria aba
        ws.cell(row, 14, f'=IF(AND(H{row}<>"",K{row}<>""),IF(H{row}=K{row},1,0),"")')

        for col in range(6, 11):
            ws.cell(row, col).fill = LLM_FILL
        for col in range(15, 20):
            ws.cell(row, col).fill = SCORE_FILL
        for col in range(1, 21):
            ws.cell(row, col).font = BASE_FONT
            ws.cell(row, col).border = BORDER

    _preparar_comparativo(wb, ultima)

    if verbose:
        acao = "criada" if nova else "atualizada"
        print(f"Aba '{ABA_LOCAL}' {acao} com {ultima - 1} linhas (gabarito copiado de '{ABA_GEMINI}').")
        print(f"Aba '{ABA_COMPARATIVO}' atualizada.")
    return ultima


def _preparar_comparativo(wb, ultima):
    if ABA_COMPARATIVO in wb.sheetnames:
        del wb[ABA_COMPARATIVO]
    ws = wb.create_sheet(ABA_COMPARATIVO)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 14

    titulos = ["Métrica", "Gemini", MODEL_LABEL, "Diferença"]
    for i, t in enumerate(titulos, start=1):
        c = ws.cell(1, i, t)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    g, l = ABA_GEMINI, ABA_LOCAL
    metricas = [
        ("Casos avaliados", f"=COUNT({g}!N2:N{ultima})", f"=COUNT({l}!N2:N{ultima})"),
        ("% Diagnóstico correto", f'=IFERROR(AVERAGE({g}!N2:N{ultima}),"")', f'=IFERROR(AVERAGE({l}!N2:N{ultima}),"")'),
        ("Média - Não revela solução", f'=IFERROR(AVERAGE({g}!O2:O{ultima}),"")', f'=IFERROR(AVERAGE({l}!O2:O{ultima}),"")'),
        ("Média - Foco em um único erro", f'=IFERROR(AVERAGE({g}!P2:P{ultima}),"")', f'=IFERROR(AVERAGE({l}!P2:P{ultima}),"")'),
        ("Média - Qualidade socrática", f'=IFERROR(AVERAGE({g}!Q2:Q{ultima}),"")', f'=IFERROR(AVERAGE({l}!Q2:Q{ultima}),"")'),
        ("Média - Clareza", f'=IFERROR(AVERAGE({g}!R2:R{ultima}),"")', f'=IFERROR(AVERAGE({l}!R2:R{ultima}),"")'),
        ("Média - Tom motivacional", f'=IFERROR(AVERAGE({g}!S2:S{ultima}),"")', f'=IFERROR(AVERAGE({l}!S2:S{ultima}),"")'),
    ]
    linha = 2
    for nome, f_gem, f_loc in metricas:
        ws.cell(linha, 1, nome).font = Font(bold=True, name="Arial", size=10)
        ws.cell(linha, 2, f_gem).font = BASE_FONT
        ws.cell(linha, 3, f_loc).font = BASE_FONT
        if nome != "Casos avaliados":
            ws.cell(linha, 4, f'=IFERROR(C{linha}-B{linha},"")').font = BASE_FONT
        linha += 1

    linha += 1
    ws.cell(linha, 1, "Concordância entre os dois modelos").font = Font(bold=True, name="Arial", size=10)
    linha += 1
    ws.cell(linha, 1, "Linhas em que ambos deram o mesmo tipo de erro").font = BASE_FONT
    ws.cell(linha, 2, f'=SUMPRODUCT(({g}!H2:H{ultima}<>"")*({l}!H2:H{ultima}<>"")*({g}!H2:H{ultima}={l}!H2:H{ultima}))').font = BASE_FONT
    linha += 1
    ws.cell(linha, 1, "% de concordância").font = BASE_FONT
    ws.cell(linha, 2, f'=IFERROR(B{linha-1}/SUMPRODUCT(({g}!H2:H{ultima}<>"")*({l}!H2:H{ultima}<>"")),"")').font = BASE_FONT

    linha += 2
    ws.cell(linha, 1, "% Diagnóstico correto por problema").font = Font(bold=True, name="Arial", size=10)
    linha += 1
    for i, t in enumerate(["Problema", "Gemini", MODEL_LABEL], start=1):
        ws.cell(linha, i, t).font = Font(bold=True, name="Arial", size=10)
    for prob in (1132, 1153, 1828, 1873):
        linha += 1
        ws.cell(linha, 1, prob).font = BASE_FONT
        ws.cell(linha, 2, f'=IFERROR(AVERAGEIF({g}!A2:A{ultima},{prob},{g}!N2:N{ultima}),"")').font = BASE_FONT
        ws.cell(linha, 3, f'=IFERROR(AVERAGEIF({l}!A2:A{ultima},{prob},{l}!N2:N{ultima}),"")').font = BASE_FONT


# ---------------------------------------------------------------------------
# Utilidades (iguais às do avaliar_lote.py)
# ---------------------------------------------------------------------------
def parse_linhas(texto):
    linhas = set()
    for parte in texto.replace(";", ",").split(","):
        parte = parte.strip()
        if not parte:
            continue
        if "-" in parte:
            inicio, _, fim = parte.partition("-")
            inicio, fim = inicio.strip(), fim.strip()
            if not (inicio.isdigit() and fim.isdigit()):
                raise ValueError(f"Intervalo inválido: '{parte}'")
            if int(inicio) > int(fim):
                raise ValueError(f"Intervalo invertido: '{parte}'")
            linhas.update(range(int(inicio), int(fim) + 1))
        else:
            if not parte.isdigit():
                raise ValueError(f"Linha inválida: '{parte}'")
            linhas.add(int(parte))
    if not linhas:
        raise ValueError("Nenhuma linha informada.")
    if min(linhas) < 2:
        raise ValueError("A linha 1 é o cabeçalho da planilha. Use linhas a partir da 2.")
    return sorted(linhas)


def escolher_modo_interativo():
    print("Como deseja avaliar (LLM local)?")
    print("  [1] Lote completo (apenas as linhas ainda não preenchidas)")
    print("  [2] Escolher uma ou mais linhas da planilha (reavaliação)")
    print("  [3] REFAZER todas as avaliações (sobrescreve tudo que já está preenchido)")
    while True:
        opcao = input("Opção [1/2/3]: ").strip()
        if opcao == "1":
            return None, {"modo": "nunca"}
        if opcao == "2":
            try:
                linhas = parse_linhas(input("Linhas (ex: 5,7,10-12): "))
            except ValueError as e:
                print(f"[ERRO] {e}")
                continue
            return linhas, {"modo": "perguntar"}
        if opcao == "3":
            print(f"Todas as colunas Local_* da aba '{ABA_LOCAL}' serão regeradas. Backup será salvo antes.")
            if input("Confirma? [s/N]: ").strip().lower() in ("s", "sim"):
                return None, {"modo": "sempre"}
            print("Cancelado.")
            continue
        print("[ERRO] Digite 1, 2 ou 3.")


def resumir(texto, limite=110):
    texto = " ".join(str(texto).split())
    return texto if len(texto) <= limite else texto[: limite - 1] + "…"


def confirmar_sobrescrita(ws, row, politica):
    if politica["modo"] == "sempre":
        return True
    if politica["modo"] == "nunca":
        return False
    print(f"\n  A linha {row} já tem avaliação:")
    print(f"    is_correct : {ws.cell(row, 6).value}")
    print(f"    tipo_erro  : {ws.cell(row, 8).value} (linha {ws.cell(row, 7).value})")
    print(f"    feedback   : {resumir(ws.cell(row, 9).value)}")
    while True:
        resposta = input("  Sobrescrever com o novo feedback? [s]im / [n]ão / [t]odas / [q] nenhuma: ").strip().lower()
        if resposta in ("s", "sim"):
            return True
        if resposta in ("n", "nao", "não"):
            return False
        if resposta in ("t", "todas"):
            politica["modo"] = "sempre"
            return True
        if resposta in ("q", "nenhuma"):
            politica["modo"] = "nunca"
            return False
        print("  [ERRO] Responda s, n, t ou q.")


def criar_backup():
    os.makedirs(BACKUP_DIR, exist_ok=True)
    nome = os.path.splitext(os.path.basename(PLANILHA_FILE))[0]
    destino = os.path.join(BACKUP_DIR, f"{nome}_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
    shutil.copy2(PLANILHA_FILE, destino)
    return destino


def normalizar_diagnostico(resultado):
    is_correct = resultado.get("is_correct") is True
    linha = resultado.get("linha_do_erro")
    tipo = resultado.get("tipo_de_erro")
    if is_correct:
        return LINHA_SEM_ERRO, TIPO_SEM_ERRO
    if tipo is None:
        tipo = "Não identificado"
    if linha is None:
        linha = LINHA_SEM_ERRO
    return linha, tipo


def processar(ws, row, politica, posicao=""):
    problem = ws.cell(row, 1).value
    student = ws.cell(row, 2).value
    attempt = ws.cell(row, 3).value
    code = ws.cell(row, 5).value

    if problem is None or code is None:
        print(f"PULANDO linha {row}: sem problema ou sem código.")
        return "pulado"

    problem_description = banco_problemas.get(str(problem))
    if not problem_description:
        print(f"[AVISO] Problema {problem} (linha {row}) não encontrado no dicionário.")
        return "pulado"

    if ws.cell(row, 6).value is not None and not confirmar_sobrescrita(ws, row, politica):
        print(f"PULANDO linha {row}: problema={problem} aluno={student} tentativa={attempt} (Já preenchido)")
        return "pulado"

    print(f"{posicao}Processando linha {row}: problema={problem} aluno={student} tentativa={attempt}...", end=" ", flush=True)

    try:
        inicio = datetime.now()
        resultado = chamar_llm_local(problem_description, code, str(problem))
        duracao = (datetime.now() - inicio).total_seconds()
        linha_erro, tipo_erro = normalizar_diagnostico(resultado)

        ws.cell(row, 6, resultado.get("is_correct"))
        ws.cell(row, 7, linha_erro)
        ws.cell(row, 8, tipo_erro)
        ws.cell(row, 9, resultado.get("feedback_formativo"))
        ws.cell(row, 10, resultado.get("dica_acao"))

        falhos = resultado.get("casos_que_falham") or []
        extra = f" (falhou nos casos de teste {falhos})" if falhos else ""
        print(f"OK! {duracao:.0f}s{extra}")
        return "ok"
    except Exception as e:
        print(f"[ERRO]: {e}")
        return "falha"


def main():
    parser = argparse.ArgumentParser(description="Avalia submissões da planilha com um LLM local (llama-server).")
    parser.add_argument("--linhas", help="Linhas da planilha a avaliar. Ex: 5,7,10-12")
    parser.add_argument("--tudo", action="store_true", help="Lote completo (só as não avaliadas), sem menu.")
    parser.add_argument("--refazer", action="store_true", help="Refaz TODAS as avaliações, sobrescrevendo.")
    parser.add_argument("--forcar", action="store_true", help="Sobrescreve sem perguntar.")
    parser.add_argument("--testar", action="store_true", help="Só testa a conexão com o servidor local.")
    parser.add_argument("--preparar", action="store_true", help="Só cria/atualiza as abas, sem avaliar.")
    args = parser.parse_args()

    if args.testar:
        testar_conexao()
        return

    if args.preparar:
        wb = openpyxl.load_workbook(PLANILHA_FILE)
        preparar_abas(wb)
        wb.save(PLANILHA_FILE)
        print(f"Planilha salva: {PLANILHA_FILE}")
        return

    if args.refazer:
        linhas_escolhidas, politica = None, {"modo": "sempre"}
    elif args.linhas:
        linhas_escolhidas = parse_linhas(args.linhas)
        politica = {"modo": "sempre" if args.forcar else "perguntar"}
    elif args.tudo:
        linhas_escolhidas, politica = None, {"modo": "sempre" if args.forcar else "nunca"}
    else:
        linhas_escolhidas, politica = escolher_modo_interativo()

    if not testar_conexao():
        return
    print()

    if politica["modo"] != "nunca":
        print(f"Backup salvo em {criar_backup()}")

    wb = openpyxl.load_workbook(PLANILHA_FILE)
    ultima = preparar_abas(wb)
    ws = wb[ABA_LOCAL]
    wb.save(PLANILHA_FILE)

    if linhas_escolhidas is None:
        linhas = list(range(2, ultima + 1))
        acao = "Refazendo TODAS as avaliações" if politica["modo"] == "sempre" else "Processamento em lote"
        print(f"\n{acao} (linhas 2 a {ultima}) na aba '{ABA_LOCAL}'...\n")
    else:
        fora = [r for r in linhas_escolhidas if r > ultima]
        if fora:
            print(f"[AVISO] Ignorando linhas fora da planilha (máx. {ultima}): {fora}")
        linhas = [r for r in linhas_escolhidas if r <= ultima]
        if not linhas:
            print("Nenhuma linha válida para processar.")
            return
        print(f"\nIniciando processamento das linhas: {linhas}\n")

    processadas, falhas = 0, 0
    for i, row in enumerate(linhas):
        status = processar(ws, row, politica, f"[{i + 1}/{len(linhas)}] ")
        if status == "pulado":
            continue
        processadas += 1
        if status == "falha":
            falhas += 1
        wb.save(PLANILHA_FILE)  # salva a cada linha: nada se perde se travar

    print(f"\nConcluído. {processadas} submissões processadas, {falhas} falhas.")
    print(f"Resultados na aba '{ABA_LOCAL}' de {PLANILHA_FILE}.")
    print(f"Compare os dois modelos na aba '{ABA_COMPARATIVO}'.")


if __name__ == "__main__":
    main()

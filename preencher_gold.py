# -*- coding: utf-8 -*-
"""Preenche as colunas Gold_* (rascunho) em rubrica_avaliacao_2.xlsx com base na
analise manual dos diffs entre tentativas consecutivas de cada aluno e casos sintéticos."""
import openpyxl

# Atualizado para o nome do arquivo mais recente
RUBRICA_FILE = "rubrica_avaliacao.xlsx"

# key: (problem, student, attempt) -> (tipo_erro, linha_erro, descricao, comentario_extra)
GOLD = {
    # ---------------- 1132 ----------------
    (1132, 1, 1): ("Sintaxe", 6, "Falta ':' no final do for (SyntaxError)", ""),
    (1132, 1, 2): ("Sintaxe", 6, "Ainda falta ':' no for -- regressao do erro de sintaxe nao corrigida", ""),
    (1132, 1, 3): ("Lógica", 7, "Nao trata o caso X > Y (falta swap dos valores antes do loop)", ""),
    (1132, 1, 4): ("Sintaxe", 3, "'int. input()' com ponto em vez de virgula -- erro de digitacao (SyntaxError)", ""),
    (1132, 1, 5): ("Sintaxe", 3, "Mesmo erro de digitacao 'int.' ainda nao corrigido", ""),
    (1132, 1, 6): ("Execução", 3, "input().split() le os dois valores da mesma linha, mas o formato de entrada tem X e Y em linhas separadas (erro ao desempacotar)", ""),
    (1132, 1, 7): ("Lógica", 7, "Nao trata o caso X > Y (mesmo problema da tentativa 3)", ""),
    (1132, 1, 8): ("Nenhum (código correto)", None, "Accepted", ""),
    (1132, 2, 1): ("Execução", 6, "'rage' deveria ser 'range' (NameError)", ""),
    (1132, 2, 2): ("Lógica", 6, "Nao trata o caso x > y (falta swap dos valores)", ""),
    (1132, 2, 3): ("Lógica", 6, "Mesmo problema: nao trata x > y", ""),
    (1132, 2, 4): ("Nenhum (código correto)", None, "Accepted", ""),
    (1132, 3, 1): ("Nenhum (código correto)", None,
                   "Unico diff para a versao aceita e trocar f-string por .format()",
                   "CASO AMBIGUO: f-string e sintaxe Python 3 valida; a diferenca sugere "
                   "incompatibilidade do juiz com uma versao antiga do Python, nao um erro real "
                   "detectavel na leitura do codigo. Considere revisar/excluir esta linha da "
                   "metrica de acuracia."),
    (1132, 3, 2): ("Nenhum (código correto)", None, "Accepted", ""),

    # ---------------- 1153 ----------------
    (1153, 1, 1): ("Nenhum (código correto)", None, "Accepted", ""),
    (1153, 2, 1): ("Lógica", 7, "'resultadoFatorial *= 1' deveria ser '*= i' -- nunca calcula o fatorial de fato", ""),
    (1153, 2, 2): ("Sintaxe", 6, "Falta ':' no final do while (SyntaxError)", ""),
    (1153, 2, 3): ("Nenhum (código correto)", None, "Accepted", ""),
    (1153, 3, 1): ("Nenhum (código correto)", None, "Accepted", ""),

    # ---------------- 1828 ----------------
    (1828, 1, 1): ("Sintaxe", 16, "Parenteses desbalanceados no print (falta um ')')", ""),
    (1828, 1, 2): ("Lógica", 3, "O texto do prompt dentro de input() e impresso na saida, incompatibilizando com o gabarito esperado", ""),
    (1828, 1, 3): ("Nenhum (código correto)", None, "Accepted", ""),
    (1828, 2, 1): ("Sintaxe", 7, "Indentacao inconsistente do bloco if (IndentationError)", ""),
    (1828, 2, 2): ("Lógica", 7, "Bloco if/print desindentado ficou fora do for -- so processa o ultimo caso", ""),
    (1828, 2, 3): ("Nenhum (código correto)", None, "Accepted", ""),
    (1828, 3, 1): ("Nenhum (código correto)", None, "Accepted", ""),

    # ---------------- 1873 ----------------
    (1873, 1, 1): ("Sintaxe", 13, "Falta ':' no final da condicao elif", ""),
    (1873, 1, 2): ("Lógica", 6, "Ordem de leitura trocada: le 'sheldon, rajesh' mas o enunciado pede rajesh antes -- inverte o vencedor", ""),
    (1873, 1, 3): ("Lógica", 20, "No segundo elif (quando rajesh vence), o codigo imprime 'sheldon' em vez de 'rajesh' (copia-e-cola)", ""),
    (1873, 1, 4): ("Lógica", 6, "Ainda nao corrigiu a ordem de leitura -- deveria ser (rajesh, sheldon)", ""),
    (1873, 1, 5): ("Execução", 6, "Usa duas chamadas de input() separadas em vez de split() -- desalinha as entradas seguintes", ""),
    (1873, 1, 6): ("Sintaxe", 6, "Parenteses desbalanceados em map() (falta um ')')", ""),
    (1873, 1, 7): ("Sintaxe", 6, "Mesmo erro de parenteses desbalanceados em map()", ""),
    (1873, 1, 8): ("Nenhum (código correto)", None, "Accepted", ""),
    (1873, 2, 1): ("Sintaxe", 11, "'de' deveria ser 'def' -- erro de digitacao na definicao da funcao", ""),
    (1873, 2, 2): ("Execução", 24, "'prit' deveria ser 'print' (NameError)", ""),
    (1873, 2, 3): ("Lógica", 5, "Inconsistencia de maiusculas/minusculas em 'Spock' no dicionario nao bate com a entrada em minusculas", ""),
    (1873, 2, 4): ("Lógica", 5, "Mesmo problema de capitalizacao de 'Spock', mesmo apos normalizar a entrada com .lower()", ""),
    (1873, 2, 5): ("Lógica", 5, "Mesmo problema de capitalizacao de 'Spock' (regrediu a normalizacao)", ""),
    (1873, 2, 6): ("Lógica", 5, "Mesmo problema de capitalizacao de 'Spock' no dicionario", ""),
    (1873, 2, 7): ("Lógica", 5, "Mesmo problema de capitalizacao de 'Spock' no dicionario",
                   "Casos 3-7 (aluno 2): mesma causa raiz persistindo -- bom exemplo para checar se o "
                   "Gemini identifica o erro repetido de forma consistente."),
    (1873, 2, 8): ("Nenhum (código correto)", None, "Accepted", ""),
    (1873, 3, 1): ("Nenhum (código correto)", None, "Accepted", ""),

    # ---------------- CASOS SINTÉTICOS (Adicionados para atingir 50 amostras) ----------------
    (1132, 901, 1): ("Lógica", 9, "Inverteu a condição do módulo (usou == ao invés de !=), somando apenas os múltiplos de 13 em vez de ignorá-los.", "Caso Sintético (Caso 44)"),
    (1132, 902, 1): ("Sintaxe", 4, "Esqueceu os dois pontos ':' no final da instrução condicional 'if x > y'.", "Caso Sintético (Caso 45)"),
    (1153, 903, 1): ("Lógica", 2, "Inicializou a variável do acumulador 'fatorial' com 0. Como a operação é de multiplicação, o resultado será sempre 0.", "Caso Sintético (Caso 46)"),
    (1153, 904, 1): ("Lógica", 4, "Definiu o 'range(1, n)' esquecendo que o limite superior no Python é exclusivo, calculando o fatorial apenas até n-1.", "Caso Sintético (Caso 47)"),
    (1828, 905, 1): ("Formatação", 7, "Esqueceu de adicionar a hashtag '#' antes do número do caso de teste no 'print()', gerando erro de apresentação.", "Caso Sintético (Caso 48) - O erro principal repete nas linhas 7, 19 e 21."),
    (1828, 906, 1): ("Execução", 4, "Usou o delimitador '.split(\",\")' em uma entrada separada por espaços, o que resulta em vetor de tamanho 1 e causa IndexError nas linhas subsequentes.", "Caso Sintético (Caso 49)"),
    (1873, 907, 1): ("Formatação", 19, "Imprimiu o nome do vencedor usando iniciais maiúsculas (ex: 'Rajesh'), enquanto o juiz automático exige caracteres totalmente em minúsculo.", "Caso Sintético (Caso 50) - Pode ser classificado também como erro de lógica/regras pelo LLM.")
}

wb = openpyxl.load_workbook(RUBRICA_FILE)
ws = wb["Avaliacao"]

filled, missing = 0, []
for row in range(2, ws.max_row + 1):
    key = (ws.cell(row, 1).value, ws.cell(row, 2).value, ws.cell(row, 3).value)
    if key not in GOLD:
        missing.append(key)
        continue
    tipo, linha, desc, extra = GOLD[key]
    ws.cell(row, 11, tipo)          # K - Gold_tipo_erro
    if linha is not None:
        ws.cell(row, 12, linha)     # L - Gold_linha_erro
    ws.cell(row, 13, f"[RASCUNHO] {desc}")  # M - Gold_descricao_correcao
    if extra:
        existing = ws.cell(row, 20).value or ""
        ws.cell(row, 20, (existing + " " if existing else "") + f"[RASCUNHO] {extra}")
    filled += 1

wb.save(RUBRICA_FILE)
print(f"Preenchidas {filled} linhas. Faltando: {missing}")